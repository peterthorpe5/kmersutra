#!/usr/bin/env python3
"""Summarise KmerSutra comparable spike-in benchmark outputs.

This script summarises a KmerSutra v0.14-style comparable benchmark run. It is
intended to replace the earlier shell script that embedded Python inline. The
script reads the comparable-run manifest, scans each sample output directory,
collates KmerSutra species calls and evidence tables, and writes publication-
ready TSV, Excel, and HTML summaries.

The script is deliberately tolerant of partial runs. If the SGE array is still
running, missing sample outputs are retained in the status and QC tables rather
than causing the summary to fail. This allows interim progress checks while the
full screen completes.

All delimited text outputs are tab-separated.
"""

from __future__ import annotations

import argparse
import html
import logging
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, Sequence

import pandas as pd

from kmersutra.benchmark_reporting import (
    is_expected_genus_neighbour,
    reporting_layer_for_call,
)

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - exercised only in minimal envs
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None


LOGGER = logging.getLogger("kmersutra_summary")

DEFAULT_POSITIVE_CALLS = (
    "present_high_confidence",
    "present_low_confidence",
    "present_in_mixed_sample",
    "mixed_species_present",
    "present",
)

DEFAULT_PANEL1_TARGETS = ("Plasmodium vivax",)
DEFAULT_BACKGROUND_CANDIDATE_CALLS = ("background_candidate_signal",)

REQUIRED_MANIFEST_COLUMNS = (
    "sample_id",
    "input_fastq",
    "benchmark_family",
    "panel",
    "replicate",
    "spike_n",
    "source_run_dir",
    "source_relative_dir",
)

NUMERIC_CALL_COLUMNS = (
    "n_unique_kmers",
    "n_positive_sequences",
    "confidence_score",
    "conflict_ratio",
    "n_k_values_positive",
    "best_k",
    "total_hits",
    "n_hits",
)

N_GENOMES_BY_FAMILY = {
    "single_genome": 1,
    "two_genome": 2,
    "three_genome": 3,
    "shuffled_negative": 1,
}


@dataclass(frozen=True)
class SummaryPaths:
    """Output paths produced by the summary workflow.

    Attributes
    ----------
    out_dir : Path
        Main summary output directory.
    sample_status : Path
        Sample-level progress/status TSV.
    calls_long : Path
        Long-form detection calls TSV.
    evidence_long : Path
        Long-form evidence TSV.
    sample_summary : Path
        One-row-per-sample summary TSV.
    progress_by_family : Path
        Progress table by benchmark family TSV.
    qc_by_family_spike : Path
        QC table by family, panel, and spike level TSV.
    performance_by_target : Path
        Tracked-target performance table TSV.
    real_world_by_sample : Path
        Observation-level real-world interpretability table TSV.
    real_world_summary : Path
        Run-level real-world interpretability summary TSV.
    off_target_summary : Path
        Off-target taxon frequency summary TSV.
    background_candidate_summary : Path
        Plausible empirical-background candidate summary TSV.
    neighbour_lineage_summary : Path
        Expected-genus neighbouring-lineage evidence summary TSV.
    runtime_summary : Path
        Runtime summary TSV.
    workbook : Path
        Formatted Excel workbook.
    html_report : Path
        HTML summary report.
    log_path : Path
        Log file path.
    """

    out_dir: Path
    sample_status: Path
    calls_long: Path
    evidence_long: Path
    sample_summary: Path
    progress_by_family: Path
    qc_by_family_spike: Path
    performance_by_target: Path
    real_world_by_sample: Path
    real_world_summary: Path
    off_target_summary: Path
    background_candidate_summary: Path
    neighbour_lineage_summary: Path
    runtime_summary: Path
    workbook: Path
    html_report: Path
    log_path: Path


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments.

    Returns
    -------
    argparse.Namespace
        Parsed arguments.
    """
    parser = argparse.ArgumentParser(
        description=(
            "Summarise KmerSutra comparable spike-in benchmark outputs and "
            "write TSV, Excel, and HTML reports."
        )
    )
    parser.add_argument(
        "--out_root",
        required=True,
        help="KmerSutra comparable benchmark output root.",
    )
    parser.add_argument(
        "--manifest",
        "--manifest_tsv",
        dest="manifest",
        default=None,
        help=(
            "Optional comparable benchmark manifest TSV. If omitted, the "
            "script auto-detects a non-empty manifest in --out_root."
        ),
    )
    parser.add_argument(
        "--out_dir",
        default=None,
        help="Summary output directory. Defaults to <out_root>/summary.",
    )
    parser.add_argument(
        "--panel1_targets",
        nargs="*",
        default=list(DEFAULT_PANEL1_TARGETS),
        help="Expected target species for single-genome benchmark rows.",
    )
    parser.add_argument(
        "--panel2_tsv",
        default=None,
        help="Optional pathogen panel 2 TSV used to infer expected targets.",
    )
    parser.add_argument(
        "--panel3_tsv",
        default=None,
        help="Optional pathogen panel 3 TSV used to infer expected targets.",
    )
    parser.add_argument(
        "--positive_calls",
        nargs="*",
        default=list(DEFAULT_POSITIVE_CALLS),
        help="KmerSutra call labels treated as positive.",
    )
    parser.add_argument(
        "--background_candidate_calls",
        nargs="*",
        default=list(DEFAULT_BACKGROUND_CANDIDATE_CALLS),
        help=(
            "Call labels treated as plausible empirical-background candidates "
            "rather than ordinary off-target species calls."
        ),
    )
    parser.add_argument(
        "--background_candidate_taxa",
        nargs="*",
        default=[],
        help=(
            "Taxon labels treated as plausible empirical-background candidates, "
            "for example 'Hammondia hammondi'."
        ),
    )
    parser.add_argument(
        "--background_candidate_file",
        default=None,
        help="Optional text file with one background-candidate taxon per line.",
    )
    parser.add_argument(
        "--demote_expected_genus_neighbours",
        action="store_true",
        help=(
            "In positive spike-in samples, count non-expected same-genus "
            "species calls as expected-lineage neighbour evidence rather than "
            "strict reportable off-target species. Raw evidence is still "
            "retained in the long-form tables."
        ),
    )
    parser.add_argument(
        "--summary_name",
        default="kmersutra_comparable_summary",
        help="Prefix used for Excel and HTML outputs.",
    )
    parser.add_argument(
        "--allow_partial",
        action="store_true",
        help="Allow missing sample outputs and summarise completed samples.",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail if any sample is missing calls or has a non-zero exit status.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print verbose logging to stderr.",
    )
    return parser.parse_args()


def configure_logging(*, log_path: Path, verbose: bool) -> None:
    """Configure console and file logging.

    Parameters
    ----------
    log_path : Path
        Path to the log file.
    verbose : bool
        Whether to emit debug-level console logging.
    """
    log_path.parent.mkdir(parents=True, exist_ok=True)
    level = logging.DEBUG if verbose else logging.INFO
    formatter = logging.Formatter(
        fmt="%(asctime)s %(levelname)s [%(name)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    LOGGER.setLevel(logging.DEBUG)
    for old_handler in LOGGER.handlers:
        old_handler.close()
    LOGGER.handlers.clear()

    console_handler = logging.StreamHandler()
    console_handler.setLevel(level)
    console_handler.setFormatter(formatter)

    file_handler = logging.FileHandler(filename=log_path, mode="w")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)

    LOGGER.addHandler(console_handler)
    LOGGER.addHandler(file_handler)


def build_summary_paths(*, out_dir: Path, summary_name: str) -> SummaryPaths:
    """Build standard output paths.

    Parameters
    ----------
    out_dir : Path
        Summary output directory.
    summary_name : str
        Output file prefix for workbook and HTML report.

    Returns
    -------
    SummaryPaths
        Standard output path bundle.
    """
    return SummaryPaths(
        out_dir=out_dir,
        sample_status=out_dir / "sample_status.tsv",
        calls_long=out_dir / "kmersutra_detection_calls_long.tsv.gz",
        evidence_long=out_dir / "kmersutra_evidence_long.tsv.gz",
        sample_summary=out_dir / "kmersutra_sample_summary.tsv",
        progress_by_family=out_dir / "progress_by_family.tsv",
        qc_by_family_spike=out_dir / "qc_by_family_spike.tsv",
        performance_by_target=out_dir / "tracked_target_performance.tsv",
        real_world_by_sample=out_dir / "real_world_by_sample.tsv",
        real_world_summary=out_dir / "real_world_summary.tsv",
        off_target_summary=out_dir / "off_target_summary.tsv",
        background_candidate_summary=out_dir / "background_candidate_summary.tsv",
        neighbour_lineage_summary=out_dir / "neighbour_lineage_summary.tsv",
        runtime_summary=out_dir / "runtime_summary.tsv",
        workbook=out_dir / f"{summary_name}.xlsx",
        html_report=out_dir / f"{summary_name}.html",
        log_path=out_dir / f"{summary_name}.log",
    )


def safe_read_tsv(*, path: Path) -> pd.DataFrame:
    """Read a TSV file safely.

    Parameters
    ----------
    path : Path
        TSV path to read.

    Returns
    -------
    pd.DataFrame
        Parsed table, or an empty table if the file is missing or unusable.
    """
    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame()
    try:
        dataframe = pd.read_csv(path, sep="\t", dtype=str)
    except Exception as exc:  # noqa: BLE001
        LOGGER.warning("Could not read TSV %s: %s", path, exc)
        return pd.DataFrame()
    dataframe.columns = [str(column).strip() for column in dataframe.columns]
    return dataframe


def write_tsv(*, dataframe: pd.DataFrame, path: Path) -> None:
    """Write a table as tab-separated text, optionally gzip-compressed.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Table to write.
    path : Path
        Output path. Paths ending ``.gz`` are written as gzip-compressed TSV.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(path, sep="\t", index=False)
    LOGGER.info("Wrote %s (%s rows)", path, len(dataframe))


def normalise_taxon_name(*, value: object) -> str:
    """Normalise a taxon name for matching.

    Parameters
    ----------
    value : object
        Raw taxon value.

    Returns
    -------
    str
        Normalised taxon name.
    """
    text = "" if value is None or pd.isna(value) else str(value)
    text = text.replace("_", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def split_label_string(*, value: object) -> list[str]:
    """Split a semicolon/comma-separated target label string.

    Parameters
    ----------
    value : object
        Raw label string.

    Returns
    -------
    list[str]
        Cleaned labels.
    """
    if value is None or pd.isna(value):
        return []
    text = str(value).replace(",", ";")
    return [normalise_taxon_name(value=part) for part in text.split(";") if part.strip()]



def read_taxon_list(*, path: Path | None) -> list[str]:
    """Read one taxon label per line from a text file.

    Parameters
    ----------
    path : pathlib.Path or None
        Optional text file. Blank lines and comment lines are ignored.

    Returns
    -------
    list[str]
        Normalised taxon labels.
    """
    if path is None:
        return []
    if not path.exists() or path.stat().st_size == 0:
        raise FileNotFoundError(f"Background-candidate taxon file missing or empty: {path}")
    labels: list[str] = []
    seen: set[str] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip() or line.strip().startswith("#"):
            continue
        label = normalise_taxon_name(value=line).lower()
        if label and label not in seen:
            seen.add(label)
            labels.append(label)
    return labels


def build_background_candidate_taxa(
    *,
    direct_taxa: Sequence[str],
    taxa_file: Path | None,
) -> set[str]:
    """Build the set of empirical-background candidate taxon labels.

    Parameters
    ----------
    direct_taxa : sequence of str
        Taxa supplied directly on the command line.
    taxa_file : pathlib.Path or None
        Optional one-label-per-line text file.

    Returns
    -------
    set[str]
        Normalised lower-case taxon labels.
    """
    labels = {
        normalise_taxon_name(value=taxon).lower()
        for taxon in direct_taxa
        if normalise_taxon_name(value=taxon)
    }
    labels.update(read_taxon_list(path=taxa_file))
    return labels


def truthy_value(*, value: object) -> bool:
    """Return True for common truthy values from TSV columns.

    Parameters
    ----------
    value : object
        Raw value.

    Returns
    -------
    bool
        Parsed truth value.
    """
    if value is None or pd.isna(value):
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y"}

def safe_identifier(*, value: str) -> str:
    """Create a safe identifier from a free-text label.

    Parameters
    ----------
    value : str
        Raw label.

    Returns
    -------
    str
        Safe identifier.
    """
    value = value.strip().replace(" ", "_")
    value = re.sub(r"[^A-Za-z0-9_.-]+", "_", value)
    value = re.sub(r"_+", "_", value)
    return value.strip("_") or "unknown"


def resolve_manifest_path(*, out_root: Path, manifest_path: Path | None = None) -> Path:
    """Resolve a comparable benchmark manifest path robustly.

    Parameters
    ----------
    out_root : Path
        KmerSutra comparable benchmark output root.
    manifest_path : Path | None
        Explicit manifest path supplied by the user. If provided, this path is
        validated and returned.

    Returns
    -------
    Path
        Resolved non-empty manifest TSV path.

    Raises
    ------
    FileNotFoundError
        If no suitable manifest file can be found.
    ValueError
        If multiple ambiguous manifest files are found.
    """
    if manifest_path is not None:
        if not manifest_path.exists() or manifest_path.stat().st_size == 0:
            raise FileNotFoundError(f"Manifest missing or empty: {manifest_path}")
        return manifest_path

    preferred_names = (
        "kmersutra_comparable_manifest.tsv",
        "kmersutra_v016_conservative_manifest.tsv",
        "kmersutra_v015_conservative_manifest.tsv",
        "kmersutra_v014_comparable_manifest.tsv",
    )
    for name in preferred_names:
        candidate = out_root / name
        if candidate.exists() and candidate.stat().st_size > 0:
            return candidate

    glob_candidates = sorted(
        candidate
        for candidate in out_root.glob("*manifest*.tsv")
        if candidate.is_file() and candidate.stat().st_size > 0
    )
    if len(glob_candidates) == 1:
        return glob_candidates[0]
    if len(glob_candidates) > 1:
        candidates = ", ".join(str(candidate) for candidate in glob_candidates)
        raise ValueError(
            "Multiple manifest-like files found. Please provide --manifest "
            f"explicitly. Candidates: {candidates}"
        )
    raise FileNotFoundError(f"No non-empty manifest TSV found in {out_root}")


def normalise_manifest_columns(*, manifest: pd.DataFrame) -> pd.DataFrame:
    """Normalise manifest column names across KmerSutra run versions.

    Parameters
    ----------
    manifest : pd.DataFrame
        Comparable benchmark manifest table loaded from TSV.

    Returns
    -------
    pd.DataFrame
        Manifest with version-specific aliases normalised to the canonical
        column names used by the summary workflow.
    """
    manifest = manifest.copy()
    column_aliases = {
        "spike_reads": "spike_n",
        "spike_reads_per_genome": "spike_n",
        "total_spiked_reads": "spike_n",
        "rep": "replicate",
        "family": "benchmark_family",
        "sample": "sample_id",
        "fastq": "input_fastq",
    }
    rename_map = {}
    for old_name, new_name in column_aliases.items():
        if old_name in manifest.columns and new_name not in manifest.columns:
            rename_map[old_name] = new_name
    if rename_map:
        LOGGER.info("Normalising manifest column aliases: %s", rename_map)
        manifest = manifest.rename(columns=rename_map)

    if "source_relative_dir" not in manifest.columns:
        if "input_fastq" in manifest.columns:
            manifest["source_relative_dir"] = manifest["input_fastq"].map(
                infer_source_relative_dir
            )
        else:
            manifest["source_relative_dir"] = ""

    return manifest


def infer_source_relative_dir(input_fastq: str) -> str:
    """Infer a source-relative sample directory from an input FASTQ path.

    Parameters
    ----------
    input_fastq : str
        Input FASTQ path from the comparable benchmark manifest.

    Returns
    -------
    str
        Parent directory name for the FASTQ, or an empty string if unavailable.
    """
    if not input_fastq:
        return ""
    return Path(str(input_fastq)).parent.name


def load_manifest(*, path: Path) -> pd.DataFrame:
    """Load and validate the comparable-run manifest.

    Parameters
    ----------
    path : Path
        Manifest TSV path.

    Returns
    -------
    pd.DataFrame
        Manifest table with required columns.

    Raises
    ------
    FileNotFoundError
        If the manifest is missing.
    ValueError
        If required columns are absent.
    """
    if not path.exists() or path.stat().st_size == 0:
        raise FileNotFoundError(f"Manifest missing or empty: {path}")
    manifest = pd.read_csv(path, sep="\t", dtype=str).fillna("")
    manifest = normalise_manifest_columns(manifest=manifest)
    missing = [column for column in REQUIRED_MANIFEST_COLUMNS if column not in manifest]
    if missing:
        raise ValueError(
            "Manifest is missing required columns after alias normalisation: "
            + ", ".join(missing)
        )
    LOGGER.info("Loaded manifest: %s (%s rows)", path, len(manifest))
    return manifest


def read_panel_targets(*, path: Path | None) -> list[str]:
    """Read expected target labels from a pathogen panel TSV.

    Parameters
    ----------
    path : Path | None
        Path to panel TSV.

    Returns
    -------
    list[str]
        Target labels found in the panel.
    """
    if path is None or not path.exists() or path.stat().st_size == 0:
        return []
    dataframe = pd.read_csv(path, sep="\t", dtype=str).fillna("")
    if dataframe.empty:
        return []

    candidate_columns = [
        "target_label",
        "species_name",
        "label",
        "name",
        "taxon_name",
    ]
    for column in candidate_columns:
        if column in dataframe.columns:
            labels = [
                normalise_taxon_name(value=value)
                for value in dataframe[column].tolist()
                if str(value).strip()
            ]
            if labels:
                LOGGER.info("Loaded %s target labels from %s", len(labels), path)
                return list(dict.fromkeys(labels))

    if dataframe.shape[1] >= 2:
        labels = [
            normalise_taxon_name(value=value)
            for value in dataframe.iloc[:, 1].tolist()
            if str(value).strip()
        ]
        LOGGER.info("Loaded %s target labels from %s second column", len(labels), path)
        return list(dict.fromkeys(labels))

    return []


def build_expected_target_map(
    *,
    panel1_targets: Sequence[str],
    panel2_tsv: Path | None,
    panel3_tsv: Path | None,
) -> dict[str, list[str]]:
    """Build expected target labels by panel.

    Parameters
    ----------
    panel1_targets : Sequence[str]
        Expected target labels for panel1/single-genome samples.
    panel2_tsv : Path | None
        Optional panel 2 TSV.
    panel3_tsv : Path | None
        Optional panel 3 TSV.

    Returns
    -------
    dict[str, list[str]]
        Expected target labels keyed by panel name.
    """
    panel1 = [normalise_taxon_name(value=value) for value in panel1_targets]
    panel2 = read_panel_targets(path=panel2_tsv)
    panel3 = read_panel_targets(path=panel3_tsv)

    mapping = {
        "panel1": panel1,
        "single": panel1,
        "shuffled": panel1,
        "panel2": panel2,
        "panel3": panel3,
    }
    LOGGER.info("Expected panel1 targets: %s", "; ".join(panel1) or "none")
    LOGGER.info("Expected panel2 targets: %s", "; ".join(panel2) or "none")
    LOGGER.info("Expected panel3 targets: %s", "; ".join(panel3) or "none")
    return mapping


def expected_targets_for_row(
    *,
    row: Mapping[str, object],
    expected_targets_by_panel: Mapping[str, Sequence[str]],
) -> list[str]:
    """Return expected targets for one manifest row.

    Parameters
    ----------
    row : Mapping[str, object]
        Manifest row.
    expected_targets_by_panel : Mapping[str, Sequence[str]]
        Expected target map keyed by panel.

    Returns
    -------
    list[str]
        Expected target labels. Shuffled samples return contextual targets but
        are still treated as negative controls by downstream metrics.
    """
    panel = str(row.get("panel", "")).strip()
    family = str(row.get("benchmark_family", "")).strip()
    if family == "two_genome":
        panel = "panel2"
    elif family == "three_genome":
        panel = "panel3"
    elif family == "single_genome":
        panel = "panel1"
    elif family == "shuffled_negative":
        panel = "shuffled"
    return list(expected_targets_by_panel.get(panel, []))


def sample_output_dir(*, out_root: Path, row: Mapping[str, object]) -> Path:
    """Resolve the sample output directory from a manifest row.

    Parameters
    ----------
    out_root : Path
        Comparable run output root.
    row : Mapping[str, object]
        Manifest row.

    Returns
    -------
    Path
        Sample output directory.
    """
    family = str(row.get("benchmark_family", "unclassified"))
    sample_id = str(row.get("sample_id", "unknown_sample"))
    nested = out_root / "samples" / family / sample_id
    if nested.exists():
        return nested
    flat = out_root / "samples" / sample_id
    if flat.exists():
        return flat
    return nested


def coerce_numeric(*, dataframe: pd.DataFrame, columns: Iterable[str]) -> pd.DataFrame:
    """Convert selected columns to numeric when present.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Input table.
    columns : Iterable[str]
        Columns to convert.

    Returns
    -------
    pd.DataFrame
        Table with numeric conversions applied.
    """
    output = dataframe.copy()
    for column in columns:
        if column in output.columns:
            output[column] = pd.to_numeric(output[column], errors="coerce")
    return output


def positive_call_mask(*, dataframe: pd.DataFrame, positive_calls: set[str]) -> pd.Series:
    """Return a positive-call mask for a KmerSutra calls table.

    Parameters
    ----------
    dataframe : pd.DataFrame
        KmerSutra calls table.
    positive_calls : set[str]
        Call labels treated as positive.

    Returns
    -------
    pd.Series
        Boolean mask indexed like the input table.
    """
    if dataframe.empty or "call" not in dataframe.columns:
        return pd.Series(False, index=dataframe.index)
    return dataframe["call"].fillna("").astype(str).isin(positive_calls)


def get_report_label(*, row: Mapping[str, object]) -> str:
    """Return the best available taxon/report label from a call row.

    Parameters
    ----------
    row : Mapping[str, object]
        KmerSutra call row.

    Returns
    -------
    str
        Report label.
    """
    for column in ("species_name", "evidence_name", "taxon_name", "target_label"):
        value = row.get(column, "")
        label = normalise_taxon_name(value=value)
        if label:
            return label
    return "unknown_taxon"


def is_species_level_row(*, row: Mapping[str, object]) -> bool:
    """Return whether a KmerSutra call row is species-level.

    Parameters
    ----------
    row : Mapping[str, object]
        KmerSutra call row.

    Returns
    -------
    bool
        True if the row appears to describe species-level evidence.
    """
    evidence_rank = str(row.get("evidence_rank", "")).strip().lower()
    if evidence_rank:
        return evidence_rank == "species"
    species_name = normalise_taxon_name(value=row.get("species_name", ""))
    return bool(species_name and species_name != "unknown_taxon")


def add_call_metadata(
    *,
    calls_df: pd.DataFrame,
    row: Mapping[str, object],
    out_root: Path,
    expected_targets: Sequence[str],
    positive_calls: set[str],
    background_candidate_calls: set[str] | None = None,
    background_candidate_taxa: set[str] | None = None,
    demote_expected_genus_neighbours: bool = False,
) -> pd.DataFrame:
    """Add manifest and classification metadata to KmerSutra call rows.

    Parameters
    ----------
    calls_df : pd.DataFrame
        Raw KmerSutra detection calls table.
    row : Mapping[str, object]
        Manifest row.
    out_root : Path
        Comparable run output root.
    expected_targets : Sequence[str]
        Expected target labels for the sample context.
    positive_calls : set[str]
        Positive KmerSutra call labels.
    demote_expected_genus_neighbours : bool, optional
        Whether to treat non-expected same-genus species in positive spike-in
        samples as expected-lineage neighbour evidence.

    Returns
    -------
    pd.DataFrame
        Annotated long-form calls table.
    """
    if calls_df.empty:
        return pd.DataFrame()

    output = calls_df.copy()
    for column in NUMERIC_CALL_COLUMNS:
        if column in output.columns:
            output[column] = pd.to_numeric(output[column], errors="coerce")

    output["report_label"] = [
        get_report_label(row=record) for record in output.to_dict(orient="records")
    ]
    output["normalised_report_label"] = output["report_label"].map(
        lambda value: normalise_taxon_name(value=value).lower()
    )
    expected_normalised = {
        normalise_taxon_name(value=target).lower() for target in expected_targets
    }
    metadata = metadata_from_manifest_row(
        row=row,
        out_root=out_root,
        expected_targets=expected_targets,
    )
    sample_is_negative = bool(metadata.get("is_negative", False))
    output["is_species_level"] = [
        is_species_level_row(row=record) for record in output.to_dict(orient="records")
    ]
    output["is_positive_call"] = positive_call_mask(
        dataframe=output,
        positive_calls=positive_calls,
    )
    background_calls = background_candidate_calls or set()
    background_taxa = background_candidate_taxa or set()
    output["is_background_candidate_call"] = output.get(
        "call",
        pd.Series("", index=output.index),
    ).fillna("").astype(str).isin(background_calls)
    if "is_background_candidate" in output.columns:
        output["is_background_candidate_column"] = output["is_background_candidate"].map(
            lambda value: truthy_value(value=value)
        )
    else:
        output["is_background_candidate_column"] = False
    output["is_background_candidate_taxon"] = output["normalised_report_label"].isin(
        background_taxa
    )
    output["is_background_candidate_signal"] = (
        output["is_background_candidate_call"]
        | output["is_background_candidate_column"]
        | output["is_background_candidate_taxon"]
    )
    output["is_expected_target"] = output["normalised_report_label"].isin(
        expected_normalised
    )
    output["is_positive_expected"] = (
        output["is_positive_call"] & output["is_expected_target"]
    )
    output["is_positive_background_candidate"] = (
        output["is_species_level"] & output["is_background_candidate_signal"]
    )
    output["is_expected_genus_neighbour"] = output.apply(
        lambda record: is_expected_genus_neighbour(
            report_label=record.get("report_label", ""),
            expected_targets=expected_targets,
            is_expected_target=bool(record.get("is_expected_target", False)),
            is_negative_sample=sample_is_negative,
            is_background_candidate=bool(
                record.get("is_background_candidate_signal", False)
            ),
            demote_expected_genus_neighbours=demote_expected_genus_neighbours,
        ),
        axis=1,
    )
    output["is_positive_neighbour_lineage"] = (
        output["is_positive_call"]
        & output["is_species_level"]
        & output["is_expected_genus_neighbour"]
    )
    output["is_positive_off_target_raw"] = (
        output["is_positive_call"] & ~output["is_expected_target"]
    )
    output["is_positive_off_target"] = (
        output["is_positive_off_target_raw"]
        & ~output["is_background_candidate_signal"]
        & ~output["is_expected_genus_neighbour"]
    )
    output["is_positive_plasmodium_off_target"] = (
        output["is_positive_off_target"]
        & output["normalised_report_label"].str.contains("plasmodium", na=False)
    )
    output["benchmark_report_layer"] = output.apply(
        lambda record: reporting_layer_for_call(
            is_positive_call=bool(record.get("is_positive_call", False)),
            is_species_level=bool(record.get("is_species_level", False)),
            is_expected_target=bool(record.get("is_expected_target", False)),
            is_background_candidate=bool(
                record.get("is_background_candidate_signal", False)
            ),
            is_expected_genus_neighbour_call=bool(
                record.get("is_expected_genus_neighbour", False)
            ),
        ),
        axis=1,
    )

    for key, value in metadata.items():
        output[key] = value

    metadata_columns = list(metadata.keys())
    remaining_columns = [column for column in output.columns if column not in metadata]
    return output[metadata_columns + remaining_columns]


def metadata_from_manifest_row(
    *,
    row: Mapping[str, object],
    out_root: Path,
    expected_targets: Sequence[str],
) -> dict[str, object]:
    """Build common sample metadata from a manifest row.

    Parameters
    ----------
    row : Mapping[str, object]
        Manifest row.
    out_root : Path
        Comparable run output root.
    expected_targets : Sequence[str]
        Expected target labels.

    Returns
    -------
    dict[str, object]
        Common sample metadata.
    """
    family = str(row.get("benchmark_family", ""))
    spike_n = pd.to_numeric(str(row.get("spike_n", "")), errors="coerce")
    replicate = pd.to_numeric(str(row.get("replicate", "")), errors="coerce")
    n_genomes = N_GENOMES_BY_FAMILY.get(family, len(expected_targets) or 1)
    is_shuffled = family == "shuffled_negative"
    is_negative = bool(is_shuffled or (pd.notna(spike_n) and float(spike_n) == 0.0))
    total_spike_n = float(spike_n) * n_genomes if pd.notna(spike_n) else math.nan
    return {
        "sample_id": str(row.get("sample_id", "")),
        "benchmark_family": family,
        "panel": str(row.get("panel", "")),
        "replicate": replicate,
        "spike_n": spike_n,
        "spike_n_per_genome": spike_n,
        "n_expected_genomes": n_genomes,
        "total_spike_n": total_spike_n,
        "is_shuffled_control": is_shuffled,
        "is_negative": is_negative,
        "expected_targets": "; ".join(expected_targets),
        "input_fastq": str(row.get("input_fastq", "")),
        "source_run_dir": str(row.get("source_run_dir", "")),
        "source_relative_dir": str(row.get("source_relative_dir", "")),
        "kmersutra_sample_out": str(sample_output_dir(out_root=out_root, row=row)),
    }


def resolve_sample_file(*, sample_dir: Path, candidate_names: Sequence[str]) -> Path:
    """Resolve a sample-level output file from version-specific names.

    Parameters
    ----------
    sample_dir : Path
        Directory containing one KmerSutra sample output.
    candidate_names : Sequence[str]
        Candidate file names in preference order.

    Returns
    -------
    Path
        First existing non-empty candidate, or the first candidate path if no
        candidate exists. Returning the first path preserves status reporting
        for missing outputs.
    """
    candidates = [sample_dir / candidate_name for candidate_name in candidate_names]
    for candidate in candidates:
        if candidate.exists() and candidate.stat().st_size > 0:
            return candidate
    return candidates[0]


def read_sample_outputs(
    *,
    out_root: Path,
    manifest: pd.DataFrame,
    expected_targets_by_panel: Mapping[str, Sequence[str]],
    positive_calls: set[str],
    background_candidate_calls: set[str] | None = None,
    background_candidate_taxa: set[str] | None = None,
    demote_expected_genus_neighbours: bool = False,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Read all sample-level KmerSutra outputs.

    Parameters
    ----------
    out_root : Path
        Comparable benchmark output root.
    manifest : pd.DataFrame
        Comparable-run manifest.
    expected_targets_by_panel : Mapping[str, Sequence[str]]
        Expected target labels by panel.
    positive_calls : set[str]
        KmerSutra call labels treated as positive.

    Returns
    -------
    tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]
        Sample status, annotated calls, annotated evidence, and timing tables.
    """
    status_rows: list[dict[str, object]] = []
    calls_frames: list[pd.DataFrame] = []
    evidence_frames: list[pd.DataFrame] = []
    timing_frames: list[pd.DataFrame] = []

    for _, manifest_row in manifest.iterrows():
        row_dict = manifest_row.to_dict()
        expected_targets = expected_targets_for_row(
            row=row_dict,
            expected_targets_by_panel=expected_targets_by_panel,
        )
        metadata = metadata_from_manifest_row(
            row=row_dict,
            out_root=out_root,
            expected_targets=expected_targets,
        )
        sample_dir = Path(metadata["kmersutra_sample_out"])
        calls_path = sample_dir / "species_detection_calls.tsv"
        evidence_path = resolve_sample_file(
            sample_dir=sample_dir,
            candidate_names=(
                "sample_species_kmer_evidence.tsv",
                "sample_taxonomic_kmer_evidence.tsv",
            ),
        )
        timing_path = sample_dir / "screen_timing.tsv"

        calls_df = safe_read_tsv(path=calls_path)
        evidence_df = safe_read_tsv(path=evidence_path)
        timing_df = safe_read_tsv(path=timing_path)

        if timing_df.empty:
            exit_status = math.nan
            runtime_seconds = math.nan
        else:
            exit_status = pd.to_numeric(
                timing_df.get("exit_status", pd.Series([math.nan])).iloc[0],
                errors="coerce",
            )
            runtime_seconds = pd.to_numeric(
                timing_df.get("runtime_seconds", pd.Series([math.nan])).iloc[0],
                errors="coerce",
            )

        if calls_df.empty:
            screen_status = "missing_calls"
        elif pd.notna(exit_status) and int(exit_status) != 0:
            screen_status = "nonzero_exit"
        else:
            screen_status = "ok"

        status_rows.append(
            {
                **metadata,
                "screen_status": screen_status,
                "calls_path": str(calls_path),
                "evidence_path": str(evidence_path),
                "timing_path": str(timing_path),
                "calls_exists": calls_path.exists(),
                "evidence_exists": evidence_path.exists(),
                "timing_exists": timing_path.exists(),
                "exit_status": exit_status,
                "runtime_seconds": runtime_seconds,
            }
        )

        if not calls_df.empty:
            calls_frames.append(
                add_call_metadata(
                    calls_df=calls_df,
                    row=row_dict,
                    out_root=out_root,
                    expected_targets=expected_targets,
                    positive_calls=positive_calls,
                    background_candidate_calls=background_candidate_calls,
                    background_candidate_taxa=background_candidate_taxa,
                    demote_expected_genus_neighbours=demote_expected_genus_neighbours,
                )
            )

        if not evidence_df.empty:
            evidence_out = evidence_df.copy()
            for key, value in metadata.items():
                evidence_out[key] = value
            metadata_columns = list(metadata.keys())
            remaining_columns = [
                column for column in evidence_out.columns if column not in metadata
            ]
            evidence_frames.append(evidence_out[metadata_columns + remaining_columns])

        if not timing_df.empty:
            timing_out = timing_df.copy()
            for key, value in metadata.items():
                timing_out[key] = value
            metadata_columns = list(metadata.keys())
            remaining_columns = [
                column for column in timing_out.columns if column not in metadata
            ]
            timing_frames.append(timing_out[metadata_columns + remaining_columns])

    status_df = pd.DataFrame(status_rows)
    calls_long = pd.concat(calls_frames, ignore_index=True) if calls_frames else pd.DataFrame()
    evidence_long = (
        pd.concat(evidence_frames, ignore_index=True) if evidence_frames else pd.DataFrame()
    )
    timing_long = pd.concat(timing_frames, ignore_index=True) if timing_frames else pd.DataFrame()
    LOGGER.info("Read %s sample status rows", len(status_df))
    LOGGER.info("Read %s detection call rows", len(calls_long))
    LOGGER.info("Read %s evidence rows", len(evidence_long))
    return status_df, calls_long, evidence_long, timing_long


def build_sample_summary(
    *,
    status_df: pd.DataFrame,
    calls_long: pd.DataFrame,
) -> pd.DataFrame:
    """Build one-row-per-sample KmerSutra summary.

    Parameters
    ----------
    status_df : pd.DataFrame
        Sample status table.
    calls_long : pd.DataFrame
        Annotated detection calls.

    Returns
    -------
    pd.DataFrame
        Sample-level summary table.
    """
    base_columns = [
        "sample_id",
        "benchmark_family",
        "panel",
        "replicate",
        "spike_n",
        "spike_n_per_genome",
        "total_spike_n",
        "n_expected_genomes",
        "is_shuffled_control",
        "is_negative",
        "expected_targets",
        "input_fastq",
        "source_run_dir",
        "source_relative_dir",
        "kmersutra_sample_out",
        "screen_status",
        "exit_status",
        "runtime_seconds",
    ]
    available_base = [column for column in base_columns if column in status_df.columns]
    summary = status_df[available_base].copy()

    if calls_long.empty:
        for column in [
            "n_positive_labels",
            "n_positive_species",
            "n_expected_species_detected",
            "n_off_target_species",
            "n_plasmodium_off_target_species",
            "n_background_candidate_species",
            "n_neighbour_lineage_species",
            "any_expected_detected",
            "all_expected_detected",
            "any_off_target_detected",
            "any_background_candidate_detected",
            "any_neighbour_lineage_detected",
            "clean_expected_positive",
            "strict_clean_expected_positive",
            "positive_labels",
            "expected_detected_labels",
            "off_target_labels",
            "plasmodium_off_target_labels",
            "background_candidate_labels",
        "neighbour_lineage_labels",
            "neighbour_lineage_labels",
        ]:
            summary[column] = 0 if column.startswith("n_") else ""
        return summary

    rows: list[dict[str, object]] = []
    grouped = calls_long.groupby("sample_id", dropna=False)
    for sample_id, group in grouped:
        positives = group.loc[group["is_positive_call"].astype(bool)].copy()
        positive_species = positives.loc[positives["is_species_level"].astype(bool)].copy()
        expected_detected = positives.loc[
            positives["is_positive_expected"].astype(bool)
        ].copy()
        off_target = positive_species.loc[
            positive_species["is_positive_off_target"].astype(bool)
        ].copy()
        plasmodium_off = positive_species.loc[
            positive_species["is_positive_plasmodium_off_target"].astype(bool)
        ].copy()
        background_candidates = positive_species.loc[
            positive_species.get(
                "is_positive_background_candidate",
                pd.Series(False, index=positive_species.index),
            ).astype(bool)
        ].copy()
        neighbour_lineage = positive_species.loc[
            positive_species.get(
                "is_positive_neighbour_lineage",
                pd.Series(False, index=positive_species.index),
            ).astype(bool)
        ].copy()
        raw_off_target_count = int(
            positive_species.loc[
                positive_species.get(
                    "is_positive_off_target_raw",
                    pd.Series(False, index=positive_species.index),
                ).astype(bool)
            ]["report_label"].nunique()
        )
        expected_targets = split_label_string(
            value=group["expected_targets"].iloc[0]
            if "expected_targets" in group.columns and not group.empty
            else ""
        )
        expected_detected_labels = sorted(
            set(expected_detected["report_label"].dropna().astype(str).tolist())
        )
        all_expected_detected = bool(
            expected_targets
            and set(normalise_taxon_name(value=value) for value in expected_targets)
            <= set(normalise_taxon_name(value=value) for value in expected_detected_labels)
        )
        rows.append(
            {
                "sample_id": sample_id,
                "n_positive_labels": int(positives["report_label"].nunique()),
                "n_positive_species": int(positive_species["report_label"].nunique()),
                "n_expected_species_detected": int(len(expected_detected_labels)),
                "n_off_target_species": int(off_target["report_label"].nunique()),
                "n_plasmodium_off_target_species": int(
                    plasmodium_off["report_label"].nunique()
                ),
                "n_background_candidate_species": int(
                    background_candidates["report_label"].nunique()
                ),
                "n_neighbour_lineage_species": int(
                    neighbour_lineage["report_label"].nunique()
                ),
                "n_raw_off_target_species_including_background": raw_off_target_count,
                "any_expected_detected": int(len(expected_detected_labels) > 0),
                "all_expected_detected": int(all_expected_detected),
                "any_off_target_detected": int(
                    off_target["report_label"].nunique() > 0
                ),
                "any_background_candidate_detected": int(
                    background_candidates["report_label"].nunique() > 0
                ),
                "any_neighbour_lineage_detected": int(
                    neighbour_lineage["report_label"].nunique() > 0
                ),
                "clean_expected_positive": int(
                    all_expected_detected and off_target["report_label"].nunique() == 0
                ),
                "strict_clean_expected_positive": int(
                    all_expected_detected and raw_off_target_count == 0
                ),
                "positive_labels": "; ".join(
                    sorted(set(positives["report_label"].dropna().astype(str)))
                ),
                "expected_detected_labels": "; ".join(expected_detected_labels),
                "off_target_labels": "; ".join(
                    sorted(set(off_target["report_label"].dropna().astype(str)))
                ),
                "plasmodium_off_target_labels": "; ".join(
                    sorted(set(plasmodium_off["report_label"].dropna().astype(str)))
                ),
                "background_candidate_labels": "; ".join(
                    sorted(set(background_candidates["report_label"].dropna().astype(str)))
                ),
                "neighbour_lineage_labels": "; ".join(
                    sorted(set(neighbour_lineage["report_label"].dropna().astype(str)))
                ),
            }
        )

    sample_metrics = pd.DataFrame(rows)
    summary = summary.merge(sample_metrics, on="sample_id", how="left")
    fill_zero = [
        "n_positive_labels",
        "n_positive_species",
        "n_expected_species_detected",
        "n_off_target_species",
        "n_plasmodium_off_target_species",
        "n_background_candidate_species",
        "n_neighbour_lineage_species",
        "n_raw_off_target_species_including_background",
        "any_expected_detected",
        "all_expected_detected",
        "any_off_target_detected",
        "any_background_candidate_detected",
        "any_neighbour_lineage_detected",
        "clean_expected_positive",
        "strict_clean_expected_positive",
    ]
    for column in fill_zero:
        if column in summary.columns:
            summary[column] = pd.to_numeric(summary[column], errors="coerce").fillna(0)
    for column in [
        "positive_labels",
        "expected_detected_labels",
        "off_target_labels",
        "plasmodium_off_target_labels",
        "background_candidate_labels",
        "neighbour_lineage_labels",
    ]:
        if column in summary.columns:
            summary[column] = summary[column].fillna("")
    return summary


def first_spike_meeting_rate(*, dataframe: pd.DataFrame, min_rate: float) -> float:
    """Return the first spike level meeting a detection-rate threshold.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Table containing ``spike_n`` and ``detected`` columns.
    min_rate : float
        Minimum detection rate.

    Returns
    -------
    float
        First spike level meeting the rate, or NaN if none do.
    """
    if dataframe.empty:
        return math.nan
    grouped = (
        dataframe.groupby("spike_n", dropna=False)["detected"]
        .mean()
        .reset_index()
        .sort_values(by="spike_n")
    )
    detected = grouped.loc[grouped["detected"] >= min_rate, "spike_n"]
    if detected.empty:
        return math.nan
    return float(detected.iloc[0])


def safe_divide(*, numerator: float, denominator: float) -> float:
    """Divide safely and return NaN for zero denominators.

    Parameters
    ----------
    numerator : float
        Numerator.
    denominator : float
        Denominator.

    Returns
    -------
    float
        Ratio, or NaN.
    """
    if denominator == 0 or pd.isna(denominator):
        return math.nan
    return float(numerator) / float(denominator)


def compute_target_performance(
    *,
    sample_summary: pd.DataFrame,
) -> pd.DataFrame:
    """Compute tracked-target performance metrics.

    Parameters
    ----------
    sample_summary : pd.DataFrame
        One-row-per-sample summary table.

    Returns
    -------
    pd.DataFrame
        Target-level performance summary.
    """
    if sample_summary.empty:
        return pd.DataFrame()

    rows: list[dict[str, object]] = []
    for (family, panel), panel_df in sample_summary.groupby(
        ["benchmark_family", "panel"],
        dropna=False,
    ):
        expected_targets = sorted(
            {
                target
                for value in panel_df["expected_targets"].dropna().astype(str)
                for target in split_label_string(value=value)
            }
        )
        for target in expected_targets:
            target_norm = normalise_taxon_name(value=target).lower()
            work = panel_df.copy()
            work["detected"] = work["expected_detected_labels"].fillna("").map(
                lambda value: target_norm
                in {
                    normalise_taxon_name(value=part).lower()
                    for part in split_label_string(value=value)
                }
            )
            positives = work.loc[~work["is_negative"].astype(bool)].copy()
            negatives = work.loc[work["is_negative"].astype(bool)].copy()
            tp = int(positives["detected"].sum())
            fn = int((~positives["detected"]).sum())
            fp = int(negatives["detected"].sum())
            tn = int((~negatives["detected"]).sum())
            sensitivity = safe_divide(numerator=tp, denominator=tp + fn)
            specificity = safe_divide(numerator=tn, denominator=tn + fp)
            precision = safe_divide(numerator=tp, denominator=tp + fp)
            f1_denominator = (2 * tp) + fp + fn
            f1_score = safe_divide(numerator=2 * tp, denominator=f1_denominator)
            rows.append(
                {
                    "benchmark_family": family,
                    "panel": panel,
                    "target_label": target,
                    "n_positive": int(positives.shape[0]),
                    "n_negative": int(negatives.shape[0]),
                    "tp": tp,
                    "fn": fn,
                    "fp": fp,
                    "tn": tn,
                    "sensitivity": sensitivity,
                    "specificity": specificity,
                    "precision": precision,
                    "f1_score": f1_score,
                    "lod50_spike_n": first_spike_meeting_rate(
                        dataframe=positives,
                        min_rate=0.50,
                    ),
                    "lod95_spike_n": first_spike_meeting_rate(
                        dataframe=positives,
                        min_rate=0.95,
                    ),
                    "lod100_spike_n": first_spike_meeting_rate(
                        dataframe=positives,
                        min_rate=1.00,
                    ),
                }
            )
    return pd.DataFrame(rows)


def build_real_world_by_sample(*, sample_summary: pd.DataFrame) -> pd.DataFrame:
    """Build observation-level real-world interpretability metrics.

    Parameters
    ----------
    sample_summary : pd.DataFrame
        One-row-per-sample summary table.

    Returns
    -------
    pd.DataFrame
        Observation-level interpretability table.
    """
    if sample_summary.empty:
        return pd.DataFrame()
    output = sample_summary.copy()
    output["any_expected_positive"] = output["any_expected_detected"].astype(int)
    output["all_expected_positive"] = output["all_expected_detected"].astype(int)
    output["clean_positive"] = output["clean_expected_positive"].astype(int)
    output["positive_with_off_target"] = output["any_off_target_detected"].astype(int)
    output["positive_with_background_candidate"] = output.get(
        "any_background_candidate_detected",
        pd.Series(0, index=output.index),
    ).astype(int)
    output["positive_with_neighbour_lineage"] = output.get(
        "any_neighbour_lineage_detected",
        pd.Series(0, index=output.index),
    ).astype(int)
    output["strict_clean_positive"] = output.get(
        "strict_clean_expected_positive",
        output["clean_expected_positive"],
    ).astype(int)
    output["taxonomic_precision"] = output.apply(
        lambda row: safe_divide(
            numerator=float(row.get("n_expected_species_detected", 0)),
            denominator=float(row.get("n_positive_species", 0)),
        ),
        axis=1,
    )
    output["negative_with_any_taxon"] = (
        output["is_negative"].astype(bool) & (output["n_positive_species"] > 0)
    ).astype(int)
    output["negative_with_off_target"] = (
        output["is_negative"].astype(bool) & (output["n_off_target_species"] > 0)
    ).astype(int)
    output["negative_with_background_candidate"] = (
        output["is_negative"].astype(bool)
        & (
            output.get(
                "n_background_candidate_species",
                pd.Series(0, index=output.index),
            )
            > 0
        )
    ).astype(int)
    return output


def summarise_real_world(*, real_world_by_sample: pd.DataFrame) -> pd.DataFrame:
    """Summarise real-world interpretability metrics by family and panel.

    Parameters
    ----------
    real_world_by_sample : pd.DataFrame
        Observation-level interpretability table.

    Returns
    -------
    pd.DataFrame
        Family/panel-level real-world summary table.
    """
    if real_world_by_sample.empty:
        return pd.DataFrame()

    rows: list[dict[str, object]] = []
    for (family, panel), group in real_world_by_sample.groupby(
        ["benchmark_family", "panel"],
        dropna=False,
    ):
        positives = group.loc[~group["is_negative"].astype(bool)].copy()
        negatives = group.loc[group["is_negative"].astype(bool)].copy()
        rows.append(
            {
                "benchmark_family": family,
                "panel": panel,
                "n_observations": int(group.shape[0]),
                "n_positive_observations": int(positives.shape[0]),
                "n_negative_observations": int(negatives.shape[0]),
                "any_expected_sensitivity": safe_divide(
                    numerator=float(positives["any_expected_positive"].sum()),
                    denominator=float(positives.shape[0]),
                ),
                "all_expected_sensitivity": safe_divide(
                    numerator=float(positives["all_expected_positive"].sum()),
                    denominator=float(positives.shape[0]),
                ),
                "clean_sensitivity": safe_divide(
                    numerator=float(positives["clean_positive"].sum()),
                    denominator=float(positives.shape[0]),
                ),
                "strict_clean_sensitivity_including_background": safe_divide(
                    numerator=float(positives["strict_clean_positive"].sum()),
                    denominator=float(positives.shape[0]),
                ),
                "positive_background_candidate_rate": safe_divide(
                    numerator=float(positives["positive_with_background_candidate"].sum()),
                    denominator=float(positives.shape[0]),
                ),
                "positive_neighbour_lineage_rate": safe_divide(
                    numerator=float(positives["positive_with_neighbour_lineage"].sum()),
                    denominator=float(positives.shape[0]),
                ),
                "positive_off_target_rate": safe_divide(
                    numerator=float(positives["positive_with_off_target"].sum()),
                    denominator=float(positives.shape[0]),
                ),
                "negative_any_taxon_rate": safe_divide(
                    numerator=float(negatives["negative_with_any_taxon"].sum()),
                    denominator=float(negatives.shape[0]),
                ),
                "negative_off_target_rate": safe_divide(
                    numerator=float(negatives["negative_with_off_target"].sum()),
                    denominator=float(negatives.shape[0]),
                ),
                "negative_background_candidate_rate": safe_divide(
                    numerator=float(negatives["negative_with_background_candidate"].sum()),
                    denominator=float(negatives.shape[0]),
                ),
                "mean_positive_species_positive_samples": (
                    float(positives["n_positive_species"].mean())
                    if not positives.empty
                    else math.nan
                ),
                "mean_off_target_species_positive_samples": (
                    float(positives["n_off_target_species"].mean())
                    if not positives.empty
                    else math.nan
                ),
                "mean_plasmodium_off_target_species_positive_samples": (
                    float(positives["n_plasmodium_off_target_species"].mean())
                    if not positives.empty
                    else math.nan
                ),
                "mean_background_candidate_species_positive_samples": (
                    float(positives.get(
                        "n_background_candidate_species",
                        pd.Series(dtype=float),
                    ).mean())
                    if not positives.empty
                    else math.nan
                ),
                "mean_neighbour_lineage_species_positive_samples": (
                    float(positives.get(
                        "n_neighbour_lineage_species",
                        pd.Series(dtype=float),
                    ).mean())
                    if not positives.empty
                    else math.nan
                ),
                "median_taxonomic_precision_positive_samples": (
                    float(positives["taxonomic_precision"].median())
                    if not positives.empty
                    else math.nan
                ),
            }
        )
    return pd.DataFrame(rows)


def summarise_progress(*, sample_summary: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build progress and QC tables.

    Parameters
    ----------
    sample_summary : pd.DataFrame
        One-row-per-sample summary table.

    Returns
    -------
    tuple[pd.DataFrame, pd.DataFrame]
        Progress-by-family table and QC-by-family-spike table.
    """
    if sample_summary.empty:
        return pd.DataFrame(), pd.DataFrame()

    progress = (
        sample_summary.groupby(
            ["benchmark_family", "panel", "screen_status"],
            dropna=False,
        )
        .agg(n_samples=("sample_id", "count"))
        .reset_index()
    )
    qc = (
        sample_summary.groupby(
            ["benchmark_family", "panel", "spike_n"],
            dropna=False,
        )
        .agg(
            n_samples=("sample_id", "count"),
            n_ok=("screen_status", lambda series: int((series == "ok").sum())),
            n_missing=("screen_status", lambda series: int((series != "ok").sum())),
            median_positive_species=("n_positive_species", "median"),
            max_positive_species=("n_positive_species", "max"),
            median_off_target_species=("n_off_target_species", "median"),
            max_off_target_species=("n_off_target_species", "max"),
            median_background_candidate_species=("n_background_candidate_species", "median"),
            max_background_candidate_species=("n_background_candidate_species", "max"),
            median_neighbour_lineage_species=("n_neighbour_lineage_species", "median"),
            max_neighbour_lineage_species=("n_neighbour_lineage_species", "max"),
        )
        .reset_index()
        .sort_values(by=["benchmark_family", "panel", "spike_n"])
    )
    return progress, qc


def summarise_off_targets(*, calls_long: pd.DataFrame) -> pd.DataFrame:
    """Summarise positive off-target calls.

    Parameters
    ----------
    calls_long : pd.DataFrame
        Annotated calls table.

    Returns
    -------
    pd.DataFrame
        Off-target summary table.
    """
    if calls_long.empty:
        return pd.DataFrame()
    subset = calls_long.loc[
        calls_long["is_positive_off_target"].astype(bool)
        & calls_long["is_species_level"].astype(bool)
    ].copy()
    if subset.empty:
        return pd.DataFrame()
    return (
        subset.groupby(
            ["benchmark_family", "panel", "report_label"],
            dropna=False,
        )
        .agg(
            n_samples=("sample_id", "nunique"),
            median_unique_kmers=("n_unique_kmers", "median")
            if "n_unique_kmers" in subset.columns
            else ("sample_id", "size"),
            median_positive_sequences=("n_positive_sequences", "median")
            if "n_positive_sequences" in subset.columns
            else ("sample_id", "size"),
        )
        .reset_index()
        .sort_values(by=["n_samples", "report_label"], ascending=[False, True])
    )



def summarise_background_candidates(*, calls_long: pd.DataFrame) -> pd.DataFrame:
    """Summarise plausible empirical-background candidate calls.

    Parameters
    ----------
    calls_long : pd.DataFrame
        Annotated calls table.

    Returns
    -------
    pd.DataFrame
        Background-candidate taxon summary table.
    """
    if calls_long.empty or "is_background_candidate_signal" not in calls_long.columns:
        return pd.DataFrame()
    subset = calls_long.loc[
        calls_long["is_background_candidate_signal"].astype(bool)
        & calls_long["is_species_level"].astype(bool)
    ].copy()
    if subset.empty:
        return pd.DataFrame()
    return (
        subset.groupby(
            ["benchmark_family", "panel", "report_label"],
            dropna=False,
        )
        .agg(
            n_samples=("sample_id", "nunique"),
            n_negative_samples=("is_negative", lambda series: int(series.astype(bool).sum())),
            n_positive_samples=("is_negative", lambda series: int((~series.astype(bool)).sum())),
            median_unique_kmers=("n_unique_kmers", "median")
            if "n_unique_kmers" in subset.columns
            else ("sample_id", "size"),
            median_positive_sequences=("n_positive_sequences", "median")
            if "n_positive_sequences" in subset.columns
            else ("sample_id", "size"),
        )
        .reset_index()
        .sort_values(by=["n_samples", "report_label"], ascending=[False, True])
    )

def summarise_neighbour_lineages(*, calls_long: pd.DataFrame) -> pd.DataFrame:
    """Summarise expected-genus neighbouring-lineage evidence.

    Parameters
    ----------
    calls_long : pd.DataFrame
        Annotated calls table.

    Returns
    -------
    pd.DataFrame
        Neighbour-lineage evidence summary table.
    """
    if calls_long.empty or "is_positive_neighbour_lineage" not in calls_long.columns:
        return pd.DataFrame()
    subset = calls_long.loc[
        calls_long["is_positive_neighbour_lineage"].astype(bool)
        & calls_long["is_species_level"].astype(bool)
    ].copy()
    if subset.empty:
        return pd.DataFrame()
    return (
        subset.groupby(
            ["benchmark_family", "panel", "report_label"],
            dropna=False,
        )
        .agg(
            n_samples=("sample_id", "nunique"),
            median_unique_kmers=("n_unique_kmers", "median")
            if "n_unique_kmers" in subset.columns
            else ("sample_id", "size"),
            median_positive_sequences=("n_positive_sequences", "median")
            if "n_positive_sequences" in subset.columns
            else ("sample_id", "size"),
            example_expected_targets=("expected_targets", "first"),
        )
        .reset_index()
        .sort_values(by=["n_samples", "report_label"], ascending=[False, True])
    )


def summarise_runtime(*, status_df: pd.DataFrame) -> pd.DataFrame:
    """Summarise runtime by benchmark family.

    Parameters
    ----------
    status_df : pd.DataFrame
        Sample status table.

    Returns
    -------
    pd.DataFrame
        Runtime summary table.
    """
    if status_df.empty or "runtime_seconds" not in status_df.columns:
        return pd.DataFrame()
    work = status_df.copy()
    work["runtime_seconds"] = pd.to_numeric(work["runtime_seconds"], errors="coerce")
    return (
        work.groupby(["benchmark_family", "panel"], dropna=False)
        .agg(
            n_samples=("sample_id", "count"),
            n_completed=("runtime_seconds", lambda series: int(series.notna().sum())),
            total_runtime_seconds=("runtime_seconds", "sum"),
            median_runtime_seconds=("runtime_seconds", "median"),
            max_runtime_seconds=("runtime_seconds", "max"),
        )
        .reset_index()
    )


def write_excel_workbook(*, tables: Mapping[str, pd.DataFrame], path: Path) -> None:
    """Write a formatted Excel workbook.

    Parameters
    ----------
    tables : Mapping[str, pd.DataFrame]
        Sheet names and tables.
    path : Path
        Output workbook path.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, dataframe in tables.items():
            safe_sheet = safe_identifier(value=sheet_name)[:31]
            dataframe.to_excel(excel_writer=writer, sheet_name=safe_sheet, index=False)
            worksheet = writer.book[safe_sheet]
            worksheet.freeze_panes = "A2"
            if worksheet.max_row >= 1 and worksheet.max_column >= 1:
                worksheet.auto_filter.ref = worksheet.dimensions
            if PatternFill is not None and Font is not None and Alignment is not None:
                header_fill = PatternFill(
                    start_color="1F4E79",
                    end_color="1F4E79",
                    fill_type="solid",
                )
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(wrap_text=True, vertical="center")
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
            if get_column_letter is not None:
                for column_cells in worksheet.columns:
                    column_letter = get_column_letter(column_cells[0].column)
                    max_length = max(
                        len(str(cell.value)) if cell.value is not None else 0
                        for cell in column_cells[:200]
                    )
                    worksheet.column_dimensions[column_letter].width = min(
                        max(max_length + 2, 10),
                        48,
                    )
    LOGGER.info("Wrote Excel workbook: %s", path)


def html_table(*, dataframe: pd.DataFrame, max_rows: int = 200) -> str:
    """Render a DataFrame as an HTML table.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Table to render.
    max_rows : int
        Maximum rows to include.

    Returns
    -------
    str
        HTML table or placeholder paragraph.
    """
    if dataframe.empty:
        return "<p class='small'>No rows available.</p>"
    view = dataframe.head(max_rows).copy()
    return view.to_html(index=False, escape=True, classes="data-table")


def write_html_report(
    *,
    tables: Mapping[str, pd.DataFrame],
    path: Path,
    out_root: Path,
    title: str,
) -> None:
    """Write an HTML summary report.

    Parameters
    ----------
    tables : Mapping[str, pd.DataFrame]
        Tables to include.
    path : Path
        Output HTML path.
    out_root : Path
        Benchmark output root.
    title : str
        Report title.
    """
    css = """
    body {
        font-family: Arial, Helvetica, sans-serif;
        margin: 28px;
        color: #1a1a1a;
    }
    h1, h2 { color: #1f4e79; }
    .note { max-width: 1100px; line-height: 1.45; }
    .small { color: #555; font-size: 13px; }
    .table-wrap {
        overflow-x: auto;
        border: 1px solid #d9e2ef;
        margin: 18px 0;
    }
    table.data-table {
        border-collapse: collapse;
        font-size: 13px;
        width: 100%;
    }
    table.data-table th {
        background: #1f4e79;
        color: white;
        padding: 7px;
        text-align: left;
        white-space: nowrap;
    }
    table.data-table td {
        border-bottom: 1px solid #e6edf5;
        padding: 6px;
        vertical-align: top;
    }
    table.data-table tr:nth-child(even) { background: #fbfdff; }
    .kpi {
        display: inline-block;
        padding: 12px 16px;
        margin: 8px 8px 8px 0;
        background: #eef5fb;
        border-left: 5px solid #1f4e79;
    }
    """
    sample_summary = tables.get("sample_summary", pd.DataFrame())
    completed = 0
    total = 0
    if not sample_summary.empty and "screen_status" in sample_summary.columns:
        total = len(sample_summary)
        completed = int((sample_summary["screen_status"] == "ok").sum())

    sections = []
    priority_order = [
        "progress_by_family",
        "qc_by_family_spike",
        "real_world_summary",
        "tracked_target_performance",
        "off_target_summary",
        "background_candidate_summary",
        "neighbour_lineage_summary",
        "runtime_summary",
        "sample_status",
    ]
    for name in priority_order:
        if name not in tables:
            continue
        sections.append(
            f"<h2>{html.escape(name.replace('_', ' ').title())}</h2>"
            "<div class='table-wrap'>"
            f"{html_table(dataframe=tables[name], max_rows=200)}"
            "</div>"
        )

    for name, dataframe in tables.items():
        if name in priority_order:
            continue
        sections.append(
            f"<h2>{html.escape(name.replace('_', ' ').title())}</h2>"
            "<div class='table-wrap'>"
            f"{html_table(dataframe=dataframe, max_rows=100)}"
            "</div>"
        )

    path.write_text(
        f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{html.escape(title)}</title>
<style>{css}</style>
</head>
<body>
<h1>{html.escape(title)}</h1>
<p class="note">
This report summarises KmerSutra comparable spike-in screening outputs. It is
safe to run while an array is still completing: unfinished samples are retained
as missing or incomplete rather than silently dropped.
</p>
<div class="kpi"><strong>Completed samples</strong><br>{completed} / {total}</div>
<div class="kpi"><strong>Output root</strong><br>{html.escape(str(out_root))}</div>
{''.join(sections)}
</body>
</html>
""",
        encoding="utf-8",
    )
    LOGGER.info("Wrote HTML report: %s", path)


def validate_completion(*, status_df: pd.DataFrame, strict: bool, allow_partial: bool) -> None:
    """Validate whether missing/non-zero sample outputs should fail the run.

    Parameters
    ----------
    status_df : pd.DataFrame
        Sample status table.
    strict : bool
        Whether to fail on any incomplete sample.
    allow_partial : bool
        Whether partial summaries are explicitly allowed.

    Raises
    ------
    RuntimeError
        If strict validation fails.
    """
    if status_df.empty or "screen_status" not in status_df.columns:
        return
    n_not_ok = int((status_df["screen_status"] != "ok").sum())
    if n_not_ok == 0:
        return
    message = f"{n_not_ok} sample(s) are incomplete or failed."
    if strict:
        raise RuntimeError(message)
    if not allow_partial:
        LOGGER.warning("%s Re-run with --strict to fail, or --allow_partial to silence.", message)
    else:
        LOGGER.info("Partial summary allowed: %s", message)


def run_summary(
    *,
    out_root: Path,
    manifest_path: Path,
    out_dir: Path,
    panel1_targets: Sequence[str],
    panel2_tsv: Path | None,
    panel3_tsv: Path | None,
    positive_calls: set[str],
    summary_name: str,
    allow_partial: bool,
    strict: bool,
    background_candidate_calls: set[str] | None = None,
    background_candidate_taxa: set[str] | None = None,
    demote_expected_genus_neighbours: bool = False,
) -> SummaryPaths:
    """Run the full summary workflow.

    Parameters
    ----------
    out_root : Path
        Comparable benchmark output root.
    manifest_path : Path
        Manifest TSV path.
    out_dir : Path
        Summary output directory.
    panel1_targets : Sequence[str]
        Expected target labels for panel1.
    panel2_tsv : Path | None
        Optional panel2 TSV.
    panel3_tsv : Path | None
        Optional panel3 TSV.
    positive_calls : set[str]
        Positive KmerSutra call labels.
    summary_name : str
        Output prefix.
    allow_partial : bool
        Allow missing sample outputs.
    strict : bool
        Fail if outputs are incomplete.

    Returns
    -------
    SummaryPaths
        Output paths written by the workflow.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    paths = build_summary_paths(out_dir=out_dir, summary_name=summary_name)
    LOGGER.info("Starting KmerSutra comparable summary")
    LOGGER.info("Output root: %s", out_root)
    LOGGER.info("Manifest: %s", manifest_path)
    LOGGER.info("Summary output directory: %s", out_dir)

    manifest = load_manifest(path=manifest_path)
    expected_targets = build_expected_target_map(
        panel1_targets=panel1_targets,
        panel2_tsv=panel2_tsv,
        panel3_tsv=panel3_tsv,
    )
    status_df, calls_long, evidence_long, timing_long = read_sample_outputs(
        out_root=out_root,
        manifest=manifest,
        expected_targets_by_panel=expected_targets,
        positive_calls=positive_calls,
        background_candidate_calls=background_candidate_calls or set(),
        background_candidate_taxa=background_candidate_taxa or set(),
        demote_expected_genus_neighbours=demote_expected_genus_neighbours,
    )
    sample_summary = build_sample_summary(
        status_df=status_df,
        calls_long=calls_long,
    )
    progress_by_family, qc_by_family_spike = summarise_progress(
        sample_summary=sample_summary,
    )
    performance_by_target = compute_target_performance(sample_summary=sample_summary)
    real_world_by_sample = build_real_world_by_sample(sample_summary=sample_summary)
    real_world_summary = summarise_real_world(
        real_world_by_sample=real_world_by_sample,
    )
    off_target_summary = summarise_off_targets(calls_long=calls_long)
    background_candidate_summary = summarise_background_candidates(calls_long=calls_long)
    neighbour_lineage_summary = summarise_neighbour_lineages(calls_long=calls_long)
    runtime_summary = summarise_runtime(status_df=status_df)

    tables = {
        "sample_status": status_df,
        "sample_summary": sample_summary,
        "progress_by_family": progress_by_family,
        "qc_by_family_spike": qc_by_family_spike,
        "tracked_target_performance": performance_by_target,
        "real_world_by_sample": real_world_by_sample,
        "real_world_summary": real_world_summary,
        "off_target_summary": off_target_summary,
        "background_candidate_summary": background_candidate_summary,
        "neighbour_lineage_summary": neighbour_lineage_summary,
        "runtime_summary": runtime_summary,
        "detection_calls_long": calls_long,
        "evidence_long": evidence_long,
        "timing_long": timing_long,
    }

    write_tsv(dataframe=status_df, path=paths.sample_status)
    write_tsv(dataframe=calls_long, path=paths.calls_long)
    write_tsv(dataframe=evidence_long, path=paths.evidence_long)
    write_tsv(dataframe=sample_summary, path=paths.sample_summary)
    write_tsv(dataframe=progress_by_family, path=paths.progress_by_family)
    write_tsv(dataframe=qc_by_family_spike, path=paths.qc_by_family_spike)
    write_tsv(dataframe=performance_by_target, path=paths.performance_by_target)
    write_tsv(dataframe=real_world_by_sample, path=paths.real_world_by_sample)
    write_tsv(dataframe=real_world_summary, path=paths.real_world_summary)
    write_tsv(dataframe=off_target_summary, path=paths.off_target_summary)
    write_tsv(dataframe=background_candidate_summary, path=paths.background_candidate_summary)
    write_tsv(dataframe=neighbour_lineage_summary, path=paths.neighbour_lineage_summary)
    write_tsv(dataframe=runtime_summary, path=paths.runtime_summary)
    write_excel_workbook(tables=tables, path=paths.workbook)
    write_html_report(
        tables=tables,
        path=paths.html_report,
        out_root=out_root,
        title="KmerSutra comparable spike-in summary",
    )
    validate_completion(
        status_df=status_df,
        strict=strict,
        allow_partial=allow_partial,
    )
    LOGGER.info("Finished KmerSutra comparable summary")
    return paths


def main() -> None:
    """Run the command-line entry point."""
    args = parse_args()
    out_root = Path(args.out_root).expanduser().resolve()
    user_manifest_path = (
        Path(args.manifest).expanduser().resolve()
        if args.manifest is not None
        else None
    )
    manifest_path = resolve_manifest_path(
        out_root=out_root,
        manifest_path=user_manifest_path,
    )
    out_dir = (
        Path(args.out_dir).expanduser().resolve()
        if args.out_dir is not None
        else out_root / "summary"
    )
    paths = build_summary_paths(out_dir=out_dir, summary_name=args.summary_name)
    configure_logging(log_path=paths.log_path, verbose=args.verbose)
    panel2_tsv = Path(args.panel2_tsv).expanduser().resolve() if args.panel2_tsv else None
    panel3_tsv = Path(args.panel3_tsv).expanduser().resolve() if args.panel3_tsv else None
    background_candidate_file = (
        Path(args.background_candidate_file).expanduser().resolve()
        if args.background_candidate_file
        else None
    )
    background_candidate_taxa = build_background_candidate_taxa(
        direct_taxa=args.background_candidate_taxa,
        taxa_file=background_candidate_file,
    )
    run_summary(
        out_root=out_root,
        manifest_path=manifest_path,
        out_dir=out_dir,
        panel1_targets=args.panel1_targets,
        panel2_tsv=panel2_tsv,
        panel3_tsv=panel3_tsv,
        positive_calls=set(args.positive_calls),
        background_candidate_calls=set(args.background_candidate_calls),
        background_candidate_taxa=background_candidate_taxa,
        demote_expected_genus_neighbours=args.demote_expected_genus_neighbours,
        summary_name=args.summary_name,
        allow_partial=args.allow_partial,
        strict=args.strict,
    )


if __name__ == "__main__":
    main()
