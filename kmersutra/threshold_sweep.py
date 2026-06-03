"""Threshold-sweep diagnostics for KmerSutra comparable benchmarks.

This module reinterprets an existing KmerSutra comparable benchmark long-call
file under several alternative rule-based promotion thresholds. It does not
rerun read screening and it does not change the underlying evidence. Instead,
it asks whether lower promotion thresholds would move observed evidence into a
reportable positive tier while preserving negative-control specificity.

All tabular outputs are tab-separated. Comma-separated outputs are not written.
"""

from __future__ import annotations

import argparse
import html
import logging
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, Sequence

import pandas as pd

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - optional reporting dependency
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None


LOGGER = logging.getLogger("kmersutra_threshold_sweep")

DEFAULT_THRESHOLD_SPECS = (
    "current_sensitive:k1,u10,p2,e10,best101,conf0.50,conflict0.10",
    "final_stress:k1,u8,p1,e8,best101,conf0.50,conflict0.10",
    "audit_u5_p1_e5:k1,u5,p1,e5,best101,conf0.50,conflict0.10",
    "audit_bestk77:k1,u8,p1,e8,best77,conf0.50,conflict0.10",
)


@dataclass(frozen=True)
class ThresholdSet:
    """Rule-based promotion threshold set.

    Attributes
    ----------
    name : str
        Short identifier used in output tables.
    min_k_values_positive : int
        Minimum number of k values with support.
    min_unique_kmers : int
        Minimum number of unique supporting markers.
    min_positive_sequences : int
        Minimum number of positive reads or sequences.
    min_exact_hits : int
        Minimum number of exact marker hits.
    min_best_k : int
        Minimum longest-supported k value.
    min_confidence_score : float
        Minimum confidence score.
    max_reportable_conflict_ratio : float
        Maximum reportable conflict ratio.
    """

    name: str
    min_k_values_positive: int
    min_unique_kmers: int
    min_positive_sequences: int
    min_exact_hits: int
    min_best_k: int
    min_confidence_score: float
    max_reportable_conflict_ratio: float


@dataclass(frozen=True)
class SweepPaths:
    """Output paths for threshold-sweep diagnostics.

    Attributes
    ----------
    out_dir : pathlib.Path
        Output directory.
    target_performance : pathlib.Path
        Target-level sensitivity/specificity table.
    real_world_summary : pathlib.Path
        Family-level interpretability table.
    off_target_summary : pathlib.Path
        Strict off-target frequency table.
    neighbour_lineage_summary : pathlib.Path
        Same-genus neighbour-lineage frequency table.
    by_spike : pathlib.Path
        Family/spike-level detection table.
    by_sample : pathlib.Path
        Sample-level threshold-sweep table.
    workbook : pathlib.Path
        Excel workbook path.
    html_report : pathlib.Path
        HTML report path.
    log_path : pathlib.Path
        Log file path.
    """

    out_dir: Path
    target_performance: Path
    real_world_summary: Path
    off_target_summary: Path
    neighbour_lineage_summary: Path
    by_spike: Path
    by_sample: Path
    workbook: Path
    html_report: Path
    log_path: Path


def configure_logging(*, log_path: Path, verbose: bool) -> None:
    """Configure console and file logging.

    Parameters
    ----------
    log_path : pathlib.Path
        Log file path.
    verbose : bool
        Whether to enable debug-level console output.
    """
    log_path.parent.mkdir(parents=True, exist_ok=True)
    formatter = logging.Formatter(
        "%(asctime)s %(levelname)s [%(name)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    level = logging.DEBUG if verbose else logging.INFO
    LOGGER.handlers.clear()
    LOGGER.setLevel(logging.DEBUG)

    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(level)
    stream_handler.setFormatter(formatter)
    LOGGER.addHandler(stream_handler)

    file_handler = logging.FileHandler(log_path, mode="w")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    LOGGER.addHandler(file_handler)


def build_paths(*, out_dir: Path, report_name: str) -> SweepPaths:
    """Build standard output paths.

    Parameters
    ----------
    out_dir : pathlib.Path
        Output directory.
    report_name : str
        HTML/Excel/log filename prefix.

    Returns
    -------
    SweepPaths
        Output path bundle.
    """
    return SweepPaths(
        out_dir=out_dir,
        target_performance=out_dir / "threshold_sweep_target_performance.tsv",
        real_world_summary=out_dir / "threshold_sweep_real_world_summary.tsv",
        off_target_summary=out_dir / "threshold_sweep_off_target_summary.tsv",
        neighbour_lineage_summary=out_dir / "threshold_sweep_neighbour_lineage_summary.tsv",
        by_spike=out_dir / "threshold_sweep_by_family_spike.tsv",
        by_sample=out_dir / "threshold_sweep_by_sample.tsv",
        workbook=out_dir / f"{report_name}.xlsx",
        html_report=out_dir / f"{report_name}.html",
        log_path=out_dir / f"{report_name}.log",
    )


def parse_bool_series(series: pd.Series) -> pd.Series:
    """Parse common boolean-like values.

    Parameters
    ----------
    series : pandas.Series
        Input values.

    Returns
    -------
    pandas.Series
        Parsed boolean values.
    """
    return series.astype(str).str.strip().str.lower().isin({"true", "1", "yes", "y"})


def safe_rate(*, numerator: int | float, denominator: int | float) -> float:
    """Calculate a rate safely.

    Parameters
    ----------
    numerator : int or float
        Numerator.
    denominator : int or float
        Denominator.

    Returns
    -------
    float
        Rate, or NaN when the denominator is zero.
    """
    if denominator == 0 or pd.isna(denominator):
        return math.nan
    return float(numerator) / float(denominator)


def split_labels(value: object) -> list[str]:
    """Split semicolon/comma-separated taxon labels.

    Parameters
    ----------
    value : object
        Raw label string.

    Returns
    -------
    list[str]
        Clean labels.
    """
    if value is None or pd.isna(value):
        return []
    text = str(value).replace(",", ";")
    return [label.strip() for label in text.split(";") if label.strip()]


def join_labels(values: Iterable[object]) -> str:
    """Join unique non-empty labels.

    Parameters
    ----------
    values : iterable
        Labels to join.

    Returns
    -------
    str
        Semicolon-separated labels.
    """
    labels = sorted({str(value).strip() for value in values if str(value).strip()})
    return "; ".join(labels)


def parse_threshold_spec(spec: str) -> ThresholdSet:
    """Parse a compact threshold specification.

    Parameters
    ----------
    spec : str
        Specification such as
        ``final_stress:k1,u8,p1,e8,best101,conf0.50,conflict0.10``.

    Returns
    -------
    ThresholdSet
        Parsed threshold set.

    Raises
    ------
    ValueError
        If the specification is malformed.
    """
    if ":" not in spec:
        raise ValueError(f"Threshold spec must contain a name and ':' separator: {spec}")
    name, raw_parts = spec.split(":", 1)
    name = name.strip()
    if not name:
        raise ValueError(f"Threshold spec has an empty name: {spec}")

    values: dict[str, float] = {}
    for raw_part in raw_parts.split(","):
        part = raw_part.strip()
        if not part:
            continue
        if part.startswith("k") and part[1:].isdigit():
            values["min_k_values_positive"] = float(part[1:])
        elif part.startswith("u") and part[1:].isdigit():
            values["min_unique_kmers"] = float(part[1:])
        elif part.startswith("p") and part[1:].isdigit():
            values["min_positive_sequences"] = float(part[1:])
        elif part.startswith("e") and part[1:].isdigit():
            values["min_exact_hits"] = float(part[1:])
        elif part.startswith("best") and part[4:].isdigit():
            values["min_best_k"] = float(part[4:])
        elif part.startswith("conflict"):
            values["max_reportable_conflict_ratio"] = float(part[len("conflict"):])
        elif part.startswith("conf"):
            values["min_confidence_score"] = float(part[len("conf"):])
        else:
            raise ValueError(f"Unrecognised threshold token '{part}' in: {spec}")

    required = {
        "min_k_values_positive",
        "min_unique_kmers",
        "min_positive_sequences",
        "min_exact_hits",
        "min_best_k",
        "min_confidence_score",
        "max_reportable_conflict_ratio",
    }
    missing = sorted(required.difference(values))
    if missing:
        raise ValueError(f"Threshold spec is missing {missing}: {spec}")

    return ThresholdSet(
        name=name,
        min_k_values_positive=int(values["min_k_values_positive"]),
        min_unique_kmers=int(values["min_unique_kmers"]),
        min_positive_sequences=int(values["min_positive_sequences"]),
        min_exact_hits=int(values["min_exact_hits"]),
        min_best_k=int(values["min_best_k"]),
        min_confidence_score=float(values["min_confidence_score"]),
        max_reportable_conflict_ratio=float(values["max_reportable_conflict_ratio"]),
    )


def read_calls_table(*, path: Path) -> pd.DataFrame:
    """Read a long detection-call table.

    Parameters
    ----------
    path : pathlib.Path
        TSV or TSV.GZ input table.

    Returns
    -------
    pandas.DataFrame
        Normalised call table.

    Raises
    ------
    FileNotFoundError
        If the file is missing or empty.
    ValueError
        If required columns are missing.
    """
    if not path.exists() or path.stat().st_size == 0:
        raise FileNotFoundError(f"Calls table missing or empty: {path}")
    calls = pd.read_csv(path, sep="\t", dtype=str).fillna("")
    required = [
        "sample_id",
        "benchmark_family",
        "panel",
        "spike_n",
        "is_negative",
        "expected_targets",
        "report_label",
        "is_species_level",
        "is_expected_target",
        "n_unique_kmers",
        "n_positive_sequences",
        "n_k_values_positive",
        "best_k",
        "n_exact_hits",
        "reportable_conflict_ratio",
        "confidence_score",
    ]
    missing = [column for column in required if column not in calls.columns]
    if missing:
        raise ValueError("Calls table is missing required columns: " + ", ".join(missing))

    numeric_columns = [
        "spike_n",
        "n_unique_kmers",
        "n_positive_sequences",
        "n_k_values_positive",
        "best_k",
        "n_exact_hits",
        "reportable_conflict_ratio",
        "confidence_score",
    ]
    for column in numeric_columns:
        calls[column] = pd.to_numeric(calls[column], errors="coerce").fillna(0)

    calls["is_negative_bool"] = parse_bool_series(calls["is_negative"])
    calls["is_species_level_bool"] = parse_bool_series(calls["is_species_level"])
    calls["is_expected_target_bool"] = parse_bool_series(calls["is_expected_target"])
    for optional in [
        "is_background_candidate_signal",
        "is_background_candidate_taxon",
        "is_expected_genus_neighbour",
        "is_positive_call",
    ]:
        if optional in calls.columns:
            calls[f"{optional}_bool"] = parse_bool_series(calls[optional])
        else:
            calls[f"{optional}_bool"] = False
    if "is_positive_call" in calls.columns:
        calls["current_positive_bool"] = calls["is_positive_call_bool"]
    else:
        calls["current_positive_bool"] = (
            calls.get("call", pd.Series("", index=calls.index))
            .astype(str)
            .str.startswith("present")
        )
    LOGGER.info("Read %s call rows from %s", len(calls), path)
    return calls


def annotate_threshold(*, calls: pd.DataFrame, threshold: ThresholdSet) -> pd.DataFrame:
    """Annotate calls under one threshold set.

    Parameters
    ----------
    calls : pandas.DataFrame
        Long call table.
    threshold : ThresholdSet
        Promotion threshold.

    Returns
    -------
    pandas.DataFrame
        Annotated table for the threshold.
    """
    output = calls.copy()
    output["threshold_name"] = threshold.name
    output["threshold_min_unique_kmers"] = threshold.min_unique_kmers
    output["threshold_min_positive_sequences"] = threshold.min_positive_sequences
    output["threshold_min_k_values_positive"] = threshold.min_k_values_positive
    output["threshold_min_exact_hits"] = threshold.min_exact_hits
    output["threshold_min_best_k"] = threshold.min_best_k
    output["threshold_min_confidence_score"] = threshold.min_confidence_score
    output["threshold_max_reportable_conflict_ratio"] = (
        threshold.max_reportable_conflict_ratio
    )
    passes_rule_threshold = (
        output["is_species_level_bool"]
        & (output["n_unique_kmers"] >= threshold.min_unique_kmers)
        & (output["n_positive_sequences"] >= threshold.min_positive_sequences)
        & (output["n_k_values_positive"] >= threshold.min_k_values_positive)
        & (output["best_k"] >= threshold.min_best_k)
        & (output["n_exact_hits"] >= threshold.min_exact_hits)
        & (output["confidence_score"] >= threshold.min_confidence_score)
        & (
            output["reportable_conflict_ratio"]
            <= threshold.max_reportable_conflict_ratio
        )
    )
    output["sweep_positive"] = output["current_positive_bool"] | passes_rule_threshold
    output["sweep_promoted_by_threshold"] = (
        output["sweep_positive"] & ~output["current_positive_bool"]
    )
    output["sweep_expected_positive"] = (
        output["sweep_positive"] & output["is_expected_target_bool"]
    )
    output["sweep_background_candidate"] = (
        output["sweep_positive"]
        & (
            output["is_background_candidate_signal_bool"]
            | output["is_background_candidate_taxon_bool"]
        )
    )
    output["sweep_neighbour_lineage"] = (
        output["sweep_positive"]
        & output["is_expected_genus_neighbour_bool"]
        & ~output["is_negative_bool"]
        & ~output["is_expected_target_bool"]
        & ~output["sweep_background_candidate"]
    )
    output["sweep_off_target"] = (
        output["sweep_positive"]
        & ~output["is_expected_target_bool"]
        & ~output["sweep_background_candidate"]
        & ~output["sweep_neighbour_lineage"]
    )
    return output


def summarise_targets(*, annotated: pd.DataFrame) -> pd.DataFrame:
    """Summarise tracked-target performance.

    Parameters
    ----------
    annotated : pandas.DataFrame
        Threshold-annotated call table.

    Returns
    -------
    pandas.DataFrame
        Target-level performance table.
    """
    expected = annotated.loc[annotated["is_expected_target_bool"]].copy()
    rows: list[dict[str, object]] = []
    for (threshold_name, family, panel, target), group in expected.groupby(
        ["threshold_name", "benchmark_family", "panel", "report_label"],
        dropna=False,
    ):
        positives = group.loc[~group["is_negative_bool"]]
        negatives = group.loc[group["is_negative_bool"]]
        positive_samples = set(positives["sample_id"])
        negative_samples = set(negatives["sample_id"])
        tp_samples = set(positives.loc[positives["sweep_expected_positive"], "sample_id"])
        fp_samples = set(negatives.loc[negatives["sweep_expected_positive"], "sample_id"])
        tp = len(tp_samples)
        fn = len(positive_samples) - tp
        fp = len(fp_samples)
        tn = len(negative_samples) - fp
        precision = safe_rate(numerator=tp, denominator=tp + fp)
        sensitivity = safe_rate(numerator=tp, denominator=tp + fn)
        rows.append(
            {
                "threshold_name": threshold_name,
                "benchmark_family": family,
                "panel": panel,
                "target_label": target,
                "n_positive": len(positive_samples),
                "n_negative": len(negative_samples),
                "tp": tp,
                "fn": fn,
                "fp": fp,
                "tn": tn,
                "sensitivity": sensitivity,
                "specificity": safe_rate(numerator=tn, denominator=tn + fp),
                "precision": precision,
                "f1_score": (
                    2 * sensitivity * precision / (sensitivity + precision)
                    if (not math.isnan(sensitivity))
                    and (not math.isnan(precision))
                    and (sensitivity + precision > 0)
                    else math.nan
                ),
                "lod50_spike_n": threshold_lod(
                    positives=positives,
                    positive_column="sweep_expected_positive",
                    proportion=0.50,
                ),
                "lod95_spike_n": threshold_lod(
                    positives=positives,
                    positive_column="sweep_expected_positive",
                    proportion=0.95,
                ),
                "lod100_spike_n": threshold_lod(
                    positives=positives,
                    positive_column="sweep_expected_positive",
                    proportion=1.00,
                ),
            }
        )
    return pd.DataFrame(rows)


def threshold_lod(*, positives: pd.DataFrame, positive_column: str, proportion: float) -> float:
    """Estimate lowest spike level reaching a target detection proportion.

    Parameters
    ----------
    positives : pandas.DataFrame
        Positive expected-target rows.
    positive_column : str
        Boolean column denoting detected samples.
    proportion : float
        Required detection proportion.

    Returns
    -------
    float
        Lowest spike level, or NaN if not reached.
    """
    if positives.empty:
        return math.nan
    work = positives.copy()
    work["spike_n"] = pd.to_numeric(work["spike_n"], errors="coerce")
    for spike_n, group in work.groupby("spike_n", dropna=True):
        if safe_rate(
            numerator=int(group[positive_column].astype(bool).sum()),
            denominator=int(group["sample_id"].nunique()),
        ) >= proportion:
            return float(spike_n)
    return math.nan


def build_sample_table(*, annotated: pd.DataFrame) -> pd.DataFrame:
    """Build threshold/sample-level interpretation table.

    Parameters
    ----------
    annotated : pandas.DataFrame
        Threshold-annotated call table.

    Returns
    -------
    pandas.DataFrame
        One row per threshold/sample.
    """
    metadata_columns = [
        "sample_id",
        "benchmark_family",
        "panel",
        "replicate",
        "spike_n",
        "spike_n_per_genome",
        "total_spike_n",
        "n_expected_genomes",
        "is_shuffled_control",
        "is_negative",
        "expected_targets",
    ]
    existing = [column for column in metadata_columns if column in annotated.columns]
    rows: list[dict[str, object]] = []
    for (threshold_name, sample_id), group in annotated.groupby(
        ["threshold_name", "sample_id"],
        dropna=False,
    ):
        base = group.iloc[0][existing].to_dict()
        expected = group.loc[group["sweep_expected_positive"]]
        backgrounds = group.loc[group["sweep_background_candidate"]]
        neighbours = group.loc[group["sweep_neighbour_lineage"]]
        off_targets = group.loc[group["sweep_off_target"]]
        expected_labels = set(split_labels(base.get("expected_targets", "")))
        expected_detected = set(expected["report_label"])
        base.update(
            {
                "threshold_name": threshold_name,
                "n_expected_detected": len(expected_detected),
                "n_positive_species": int(group.loc[group["sweep_positive"], "report_label"].nunique()),
                "n_background_candidate_species": int(backgrounds["report_label"].nunique()),
                "n_neighbour_lineage_species": int(neighbours["report_label"].nunique()),
                "n_off_target_species": int(off_targets["report_label"].nunique()),
                "any_expected_detected": int(bool(expected_detected)),
                "all_expected_detected": int(bool(expected_labels) and expected_labels <= expected_detected),
                "clean_expected_positive": int(bool(expected_detected) and off_targets.empty),
                "strict_clean_expected_positive": int(bool(expected_detected) and off_targets.empty and backgrounds.empty),
                "expected_detected_labels": join_labels(expected["report_label"]),
                "background_candidate_labels": join_labels(backgrounds["report_label"]),
                "neighbour_lineage_labels": join_labels(neighbours["report_label"]),
                "off_target_labels": join_labels(off_targets["report_label"]),
            }
        )
        rows.append(base)
    return pd.DataFrame(rows)


def summarise_real_world(*, by_sample: pd.DataFrame) -> pd.DataFrame:
    """Summarise family-level interpretability metrics.

    Parameters
    ----------
    by_sample : pandas.DataFrame
        Sample-level threshold-sweep table.

    Returns
    -------
    pandas.DataFrame
        Family-level summary.
    """
    rows: list[dict[str, object]] = []
    for (threshold_name, family, panel), group in by_sample.groupby(
        ["threshold_name", "benchmark_family", "panel"],
        dropna=False,
    ):
        is_negative = parse_bool_series(group["is_negative"])
        positives = group.loc[~is_negative]
        negatives = group.loc[is_negative]
        rows.append(
            {
                "threshold_name": threshold_name,
                "benchmark_family": family,
                "panel": panel,
                "n_observations": int(group.shape[0]),
                "n_positive_observations": int(positives.shape[0]),
                "n_negative_observations": int(negatives.shape[0]),
                "any_expected_sensitivity": safe_rate(
                    numerator=int(positives["any_expected_detected"].sum()),
                    denominator=int(positives.shape[0]),
                ),
                "all_expected_sensitivity": safe_rate(
                    numerator=int(positives["all_expected_detected"].sum()),
                    denominator=int(positives.shape[0]),
                ),
                "clean_sensitivity": safe_rate(
                    numerator=int(positives["clean_expected_positive"].sum()),
                    denominator=int(positives.shape[0]),
                ),
                "strict_clean_sensitivity_including_background": safe_rate(
                    numerator=int(positives["strict_clean_expected_positive"].sum()),
                    denominator=int(positives.shape[0]),
                ),
                "positive_background_candidate_rate": safe_rate(
                    numerator=int((positives["n_background_candidate_species"] > 0).sum()),
                    denominator=int(positives.shape[0]),
                ),
                "positive_neighbour_lineage_rate": safe_rate(
                    numerator=int((positives["n_neighbour_lineage_species"] > 0).sum()),
                    denominator=int(positives.shape[0]),
                ),
                "positive_off_target_rate": safe_rate(
                    numerator=int((positives["n_off_target_species"] > 0).sum()),
                    denominator=int(positives.shape[0]),
                ),
                "negative_any_taxon_rate": safe_rate(
                    numerator=int((negatives["n_positive_species"] > 0).sum()),
                    denominator=int(negatives.shape[0]),
                ),
                "negative_off_target_rate": safe_rate(
                    numerator=int((negatives["n_off_target_species"] > 0).sum()),
                    denominator=int(negatives.shape[0]),
                ),
                "negative_background_candidate_rate": safe_rate(
                    numerator=int((negatives["n_background_candidate_species"] > 0).sum()),
                    denominator=int(negatives.shape[0]),
                ),
                "mean_neighbour_lineage_species_positive_samples": (
                    float(positives["n_neighbour_lineage_species"].mean())
                    if not positives.empty
                    else math.nan
                ),
                "mean_off_target_species_positive_samples": (
                    float(positives["n_off_target_species"].mean())
                    if not positives.empty
                    else math.nan
                ),
            }
        )
    return pd.DataFrame(rows)


def summarise_by_spike(*, by_sample: pd.DataFrame) -> pd.DataFrame:
    """Summarise threshold results by benchmark family and spike level.

    Parameters
    ----------
    by_sample : pandas.DataFrame
        Sample-level threshold-sweep table.

    Returns
    -------
    pandas.DataFrame
        Family/spike-level summary.
    """
    return (
        by_sample.groupby(["threshold_name", "benchmark_family", "panel", "spike_n"], dropna=False)
        .agg(
            n_samples=("sample_id", "count"),
            any_expected_rate=("any_expected_detected", "mean"),
            all_expected_rate=("all_expected_detected", "mean"),
            clean_expected_rate=("clean_expected_positive", "mean"),
            mean_off_target_species=("n_off_target_species", "mean"),
            max_off_target_species=("n_off_target_species", "max"),
            mean_neighbour_lineage_species=("n_neighbour_lineage_species", "mean"),
            max_neighbour_lineage_species=("n_neighbour_lineage_species", "max"),
        )
        .reset_index()
        .sort_values(["threshold_name", "benchmark_family", "panel", "spike_n"])
    )


def summarise_positive_taxa(
    *, annotated: pd.DataFrame, flag_column: str, table_name: str
) -> pd.DataFrame:
    """Summarise positive taxa for a threshold flag.

    Parameters
    ----------
    annotated : pandas.DataFrame
        Threshold-annotated call table.
    flag_column : str
        Boolean column defining taxa to summarise.
    table_name : str
        Name used in warning messages.

    Returns
    -------
    pandas.DataFrame
        Taxon frequency table.
    """
    subset = annotated.loc[annotated[flag_column].astype(bool)].copy()
    if subset.empty:
        LOGGER.info("No rows available for %s", table_name)
        return pd.DataFrame(
            columns=[
                "threshold_name",
                "benchmark_family",
                "panel",
                "report_label",
                "n_samples",
                "n_negative_samples",
                "n_positive_samples",
                "median_unique_kmers",
                "median_positive_sequences",
            ]
        )
    return (
        subset.groupby(["threshold_name", "benchmark_family", "panel", "report_label"], dropna=False)
        .agg(
            n_samples=("sample_id", "nunique"),
            n_negative_samples=("is_negative_bool", lambda values: int(values.astype(bool).sum())),
            n_positive_samples=("is_negative_bool", lambda values: int((~values.astype(bool)).sum())),
            median_unique_kmers=("n_unique_kmers", "median"),
            median_positive_sequences=("n_positive_sequences", "median"),
        )
        .reset_index()
        .sort_values(["threshold_name", "n_samples", "report_label"], ascending=[True, False, True])
    )


def write_tsv(*, dataframe: pd.DataFrame, path: Path) -> None:
    """Write a tab-separated table.

    Parameters
    ----------
    dataframe : pandas.DataFrame
        Table to write.
    path : pathlib.Path
        Output path.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(path, sep="\t", index=False)
    LOGGER.info("Wrote %s (%s rows)", path, len(dataframe))


def write_excel(*, tables: Mapping[str, pd.DataFrame], path: Path) -> None:
    """Write an Excel workbook with formatted headers.

    Parameters
    ----------
    tables : mapping
        Sheet names and DataFrames.
    path : pathlib.Path
        Workbook path.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, dataframe in tables.items():
            sheet = name[:31]
            dataframe.to_excel(writer, sheet_name=sheet, index=False)
            worksheet = writer.book[sheet]
            worksheet.freeze_panes = "A2"
            if worksheet.max_row >= 1 and worksheet.max_column >= 1:
                worksheet.auto_filter.ref = worksheet.dimensions
            if PatternFill is not None and Font is not None and Alignment is not None:
                header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                header_alignment = Alignment(wrap_text=True, vertical="center")
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
            if get_column_letter is not None:
                for column_cells in worksheet.columns:
                    column_letter = get_column_letter(column_cells[0].column)
                    max_length = max(
                        len(str(cell.value)) if cell.value is not None else 0
                        for cell in column_cells[:200]
                    )
                    worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 48)
    LOGGER.info("Wrote %s", path)


def html_table(dataframe: pd.DataFrame, *, max_rows: int = 200) -> str:
    """Render a DataFrame as HTML.

    Parameters
    ----------
    dataframe : pandas.DataFrame
        Table to render.
    max_rows : int
        Maximum number of rows.

    Returns
    -------
    str
        HTML table or placeholder.
    """
    if dataframe.empty:
        return "<p>No rows available.</p>"
    return dataframe.head(max_rows).to_html(index=False, escape=True)


def write_html(*, tables: Mapping[str, pd.DataFrame], path: Path) -> None:
    """Write an HTML threshold-sweep report.

    Parameters
    ----------
    tables : mapping
        Tables to include.
    path : pathlib.Path
        HTML output path.
    """
    sections = []
    for name, dataframe in tables.items():
        sections.append(
            f"<h2>{html.escape(name.replace('_', ' ').title())}</h2>"
            f"<div class='table-wrap'>{html_table(dataframe, max_rows=200)}</div>"
        )
    path.write_text(
        """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>KmerSutra threshold sweep</title>
<style>
body { font-family: Arial, Helvetica, sans-serif; margin: 28px; }
h1, h2 { color: #1f4e79; }
.table-wrap { overflow-x: auto; border: 1px solid #d9e2ef; margin: 18px 0; }
table { border-collapse: collapse; font-size: 13px; width: 100%; }
th { background: #1f4e79; color: white; padding: 7px; text-align: left; }
td { border-bottom: 1px solid #e6edf5; padding: 6px; }
</style>
</head>
<body>
<h1>KmerSutra threshold sweep</h1>
<p>This report reinterprets existing long-call evidence under alternative
rule-based promotion thresholds. It does not rerun screening.</p>
"""
        + "".join(sections)
        + "</body>\n</html>\n",
        encoding="utf-8",
    )
    LOGGER.info("Wrote %s", path)


def run_threshold_sweep(
    *, calls_table: Path, out_dir: Path, threshold_specs: Sequence[str], report_name: str
) -> SweepPaths:
    """Run threshold-sweep diagnostics.

    Parameters
    ----------
    calls_table : pathlib.Path
        Long detection-call TSV or TSV.GZ table.
    out_dir : pathlib.Path
        Output directory.
    threshold_specs : sequence of str
        Compact threshold specifications.
    report_name : str
        Excel/HTML/log prefix.

    Returns
    -------
    SweepPaths
        Output path bundle.
    """
    paths = build_paths(out_dir=out_dir, report_name=report_name)
    thresholds = [parse_threshold_spec(spec) for spec in threshold_specs]
    calls = read_calls_table(path=calls_table)
    annotated = pd.concat(
        [annotate_threshold(calls=calls, threshold=threshold) for threshold in thresholds],
        ignore_index=True,
    )
    target_performance = summarise_targets(annotated=annotated)
    by_sample = build_sample_table(annotated=annotated)
    real_world = summarise_real_world(by_sample=by_sample)
    by_spike = summarise_by_spike(by_sample=by_sample)
    off_targets = summarise_positive_taxa(
        annotated=annotated,
        flag_column="sweep_off_target",
        table_name="strict off-targets",
    )
    neighbours = summarise_positive_taxa(
        annotated=annotated,
        flag_column="sweep_neighbour_lineage",
        table_name="neighbour lineages",
    )

    tables = {
        "target_performance": target_performance,
        "real_world_summary": real_world,
        "by_family_spike": by_spike,
        "off_target_summary": off_targets,
        "neighbour_lineage_summary": neighbours,
        "by_sample": by_sample,
    }
    write_tsv(dataframe=target_performance, path=paths.target_performance)
    write_tsv(dataframe=real_world, path=paths.real_world_summary)
    write_tsv(dataframe=by_spike, path=paths.by_spike)
    write_tsv(dataframe=off_targets, path=paths.off_target_summary)
    write_tsv(dataframe=neighbours, path=paths.neighbour_lineage_summary)
    write_tsv(dataframe=by_sample, path=paths.by_sample)
    write_excel(tables=tables, path=paths.workbook)
    write_html(tables=tables, path=paths.html_report)
    return paths


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    """Parse command-line arguments.

    Parameters
    ----------
    argv : sequence of str, optional
        Argument list for tests. Defaults to ``sys.argv``.

    Returns
    -------
    argparse.Namespace
        Parsed arguments.
    """
    parser = argparse.ArgumentParser(
        description="Reinterpret KmerSutra long-call tables under alternative thresholds."
    )
    parser.add_argument(
        "--calls_table",
        required=True,
        help="Long detection-call table from comparable summary, TSV or TSV.GZ.",
    )
    parser.add_argument(
        "--out_dir",
        required=True,
        help="Output directory for threshold-sweep diagnostics.",
    )
    parser.add_argument(
        "--threshold",
        dest="thresholds",
        action="append",
        default=None,
        help=(
            "Threshold spec. May be supplied multiple times. Example: "
            "final_stress:k1,u8,p1,e8,best101,conf0.50,conflict0.10"
        ),
    )
    parser.add_argument(
        "--report_name",
        default="kmersutra_threshold_sweep",
        help="Prefix for Excel, HTML and log outputs.",
    )
    parser.add_argument("--verbose", action="store_true", help="Enable verbose logging.")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> None:
    """Run the threshold-sweep command-line entry point.

    Parameters
    ----------
    argv : sequence of str, optional
        Argument list for tests. Defaults to ``sys.argv``.
    """
    args = parse_args(argv=argv)
    out_dir = Path(args.out_dir).expanduser().resolve()
    paths = build_paths(out_dir=out_dir, report_name=args.report_name)
    configure_logging(log_path=paths.log_path, verbose=args.verbose)
    run_threshold_sweep(
        calls_table=Path(args.calls_table).expanduser().resolve(),
        out_dir=out_dir,
        threshold_specs=args.thresholds or DEFAULT_THRESHOLD_SPECS,
        report_name=args.report_name,
    )
    LOGGER.info("Finished KmerSutra threshold sweep")


if __name__ == "__main__":  # pragma: no cover
    main()
