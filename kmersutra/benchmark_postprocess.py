"""Benchmark post-processing helpers for KmerSutra.

This module intentionally operates after database building and sample screening.
It combines comparable-summary call tables, LCA reporting and optional AI
call-calibration preparation into one reproducible post-processing bundle. The
build step is kept separate because AI training needs benchmark/screening
outputs rather than the reference database alone.
"""

from __future__ import annotations

import html
import logging
from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path

from kmersutra.ai_calibration import (
    train_evaluate_call_calibrator,
    write_call_training_table_from_table,
)
from kmersutra.lca_reporting import summarise_lca_table
from kmersutra.table_io import read_records_table

DEFAULT_CALL_TABLE_NAMES = [
    "kmersutra_detection_calls_long.tsv.gz",
    "kmersutra_detection_calls_long.tsv",
    "kmersutra_detection_calls_long.parquet",
]
DEFAULT_LCA_TABLE_NAME = "kmersutra_lca_summary.tsv.gz"
DEFAULT_TRAINING_TABLE_NAME = "ai_call_training.tsv.gz"
DEFAULT_AI_MODEL_NAME = "kmersutra_call_calibrator.json"
DEFAULT_AI_SUMMARY_NAME = "kmersutra_call_calibrator_training_summary.tsv.gz"
DEFAULT_AI_EVALUATION_NAME = "kmersutra_call_calibrator_evaluation.tsv.gz"
DEFAULT_REPORT_STEM = "kmersutra_benchmark_postprocess"
REPORT_PREVIEW_ROWS = 5000


class BenchmarkPostprocessError(RuntimeError):
    """Raised when benchmark post-processing inputs are incomplete."""


def resolve_existing_table(
    *,
    explicit_path: str | Path | None,
    search_dir: str | Path | None,
    candidate_names: Sequence[str],
    description: str,
) -> Path:
    """Resolve an existing benchmark table.

    Parameters
    ----------
    explicit_path : str or pathlib.Path or None
        User-provided path. If supplied, this path must exist and be non-empty.
    search_dir : str or pathlib.Path or None
        Directory searched when ``explicit_path`` is not supplied.
    candidate_names : sequence of str
        Candidate file names searched in order.
    description : str
        Human-readable table description for error messages.

    Returns
    -------
    pathlib.Path
        Resolved existing table path.

    Raises
    ------
    BenchmarkPostprocessError
        If no suitable table can be found.
    """
    if explicit_path is not None:
        path = Path(explicit_path).expanduser().resolve()
        if path.exists() and path.stat().st_size > 0:
            return path
        raise BenchmarkPostprocessError(
            f"{description} was supplied but does not exist or is empty: {path}"
        )
    if search_dir is None:
        raise BenchmarkPostprocessError(
            f"No {description} path supplied and no search directory was provided"
        )
    root = Path(search_dir).expanduser().resolve()
    for name in candidate_names:
        path = root / name
        if path.exists() and path.stat().st_size > 0:
            return path
    candidates = ", ".join(candidate_names)
    raise BenchmarkPostprocessError(
        f"Could not find {description} in {root}. Tried: {candidates}"
    )


def _normalise_value(*, value: object) -> str:
    """Return a safe string for reports.

    Parameters
    ----------
    value : object
        Input value.

    Returns
    -------
    str
        String representation with blanks preserved as empty strings.
    """
    return "" if value is None else str(value)


def _to_float(*, value: object, default: float = 0.0) -> float:
    """Parse a numeric value defensively.

    Parameters
    ----------
    value : object
        Input value.
    default : float, optional
        Value returned when parsing fails.

    Returns
    -------
    float
        Parsed numeric value.
    """
    if value in (None, ""):
        return default
    try:
        return float(str(value))
    except (TypeError, ValueError):
        return default


def count_records_by_column(
    *,
    records: Iterable[Mapping[str, object]],
    column: str,
    output_count_column: str = "n_records",
) -> list[dict[str, object]]:
    """Count records by one table column.

    Parameters
    ----------
    records : iterable of mappings
        Input records.
    column : str
        Column to count.
    output_count_column : str, optional
        Count-column name.

    Returns
    -------
    list[dict[str, object]]
        Count records sorted by descending count and then label.
    """
    counts: Counter[str] = Counter()
    for record in records:
        value = _normalise_value(value=record.get(column, "")) or "<blank>"
        counts[value] += 1
    rows = [
        {column: value, output_count_column: count}
        for value, count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    ]
    return rows


def summarise_lca_scopes(
    *,
    lca_records: Iterable[Mapping[str, object]],
) -> list[dict[str, object]]:
    """Summarise LCA assignments by scope, rank and interpretation.

    Parameters
    ----------
    lca_records : iterable of mappings
        Records from ``kmersutra-summarise-lca``.

    Returns
    -------
    list[dict[str, object]]
        Summary rows.
    """
    grouped: dict[tuple[str, str, str, str], dict[str, object]] = {}
    for record in lca_records:
        key = (
            _normalise_value(value=record.get("lca_scope", "<blank>")),
            _normalise_value(value=record.get("lca_rank", "<blank>")),
            _normalise_value(value=record.get("lca_interpretation", "<blank>")),
            _normalise_value(value=record.get("lca_name", "<blank>")),
        )
        row = grouped.setdefault(
            key,
            {
                "lca_scope": key[0],
                "lca_rank": key[1],
                "lca_interpretation": key[2],
                "lca_name": key[3],
                "n_samples_or_scope_rows": 0,
                "sum_unique_kmers": 0.0,
                "sum_positive_sequences": 0.0,
                "max_best_k": 0.0,
            },
        )
        row["n_samples_or_scope_rows"] = int(row["n_samples_or_scope_rows"]) + 1
        row["sum_unique_kmers"] = float(row["sum_unique_kmers"]) + _to_float(
            value=record.get("total_unique_kmers", 0.0)
        )
        row["sum_positive_sequences"] = float(row["sum_positive_sequences"]) + _to_float(
            value=record.get("total_positive_sequences", 0.0)
        )
        row["max_best_k"] = max(
            float(row["max_best_k"]),
            _to_float(value=record.get("max_best_k", 0.0)),
        )
    return sorted(
        grouped.values(),
        key=lambda row: (
            str(row["lca_scope"]),
            -int(row["n_samples_or_scope_rows"]),
            str(row["lca_name"]),
        ),
    )


def summarise_numeric_features(
    *,
    records: Iterable[Mapping[str, object]],
    feature_prefixes: Sequence[str] = ("lca_",),
) -> list[dict[str, object]]:
    """Summarise numeric feature columns in AI-ready records.

    Parameters
    ----------
    records : iterable of mappings
        AI training records.
    feature_prefixes : sequence of str, optional
        Prefixes used to select feature columns.

    Returns
    -------
    list[dict[str, object]]
        Feature summary rows.
    """
    rows = [dict(record) for record in records]
    if not rows:
        return []
    columns = sorted(
        column
        for column in rows[0]
        if any(str(column).startswith(prefix) for prefix in feature_prefixes)
    )
    summary: list[dict[str, object]] = []
    for column in columns:
        values = [_to_float(value=row.get(column, 0.0)) for row in rows]
        if not values:
            continue
        non_zero = sum(1 for value in values if value != 0.0)
        summary.append(
            {
                "feature": column,
                "n_records": len(values),
                "n_non_zero": non_zero,
                "min": min(values),
                "mean": sum(values) / len(values),
                "max": max(values),
            }
        )
    return summary


def _escape_cell(*, value: object) -> str:
    """Return an HTML-escaped table cell string.

    Parameters
    ----------
    value : object
        Cell value.

    Returns
    -------
    str
        Escaped text.
    """
    return html.escape(_normalise_value(value=value))


def html_table(
    *,
    records: Sequence[Mapping[str, object]],
    max_rows: int = REPORT_PREVIEW_ROWS,
) -> str:
    """Render records as a compact HTML table.

    Parameters
    ----------
    records : sequence of mappings
        Records to render.
    max_rows : int, optional
        Maximum number of rows to include.

    Returns
    -------
    str
        HTML table or placeholder paragraph.
    """
    rows = [dict(record) for record in records]
    if not rows:
        return "<p class='small'>No rows available.</p>"
    fieldnames = list(rows[0].keys())
    header = "".join(f"<th>{html.escape(field)}</th>" for field in fieldnames)
    body_rows = []
    for record in rows[:max_rows]:
        cells = "".join(f"<td>{_escape_cell(value=record.get(field, ''))}</td>" for field in fieldnames)
        body_rows.append(f"<tr>{cells}</tr>")
    note = ""
    if len(rows) > max_rows:
        note = f"<p class='small'>Showing {max_rows} of {len(rows)} rows.</p>"
    return f"{note}<table class='data-table'><thead><tr>{header}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"


def write_html_report(
    *,
    tables: Mapping[str, Sequence[Mapping[str, object]]],
    output_path: str | Path,
    metadata: Mapping[str, object],
    title: str = "KmerSutra benchmark post-processing report",
    max_rows: int = REPORT_PREVIEW_ROWS,
    logger: logging.Logger | None = None,
) -> None:
    """Write a benchmark post-processing HTML report.

    Parameters
    ----------
    tables : mapping
        Table name to records.
    output_path : str or pathlib.Path
        HTML output path.
    metadata : mapping
        Report metadata to show as key-value cards.
    title : str, optional
        Report title.
    max_rows : int, optional
        Maximum rows per table in HTML previews.
    logger : logging.Logger or None, optional
        Logger.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    css = """
    body { font-family: Arial, Helvetica, sans-serif; margin: 28px; color: #1a1a1a; }
    h1, h2 { color: #1f4e79; }
    .note { max-width: 1100px; line-height: 1.45; }
    .small { color: #555; font-size: 13px; }
    .kpi { display: inline-block; padding: 12px 16px; margin: 8px 8px 8px 0;
           background: #eef5fb; border-left: 5px solid #1f4e79; max-width: 520px; }
    .table-wrap { overflow-x: auto; border: 1px solid #d9e2ef; margin: 18px 0; }
    table.data-table { border-collapse: collapse; font-size: 13px; width: 100%; }
    table.data-table th { background: #1f4e79; color: white; padding: 7px;
                          text-align: left; white-space: nowrap; }
    table.data-table td { border-bottom: 1px solid #e6edf5; padding: 6px; vertical-align: top; }
    table.data-table tr:nth-child(even) { background: #fbfdff; }
    """
    cards = "".join(
        "<div class='kpi'><strong>"
        f"{html.escape(str(key).replace('_', ' ').title())}</strong><br>"
        f"{html.escape(str(value))}</div>"
        for key, value in metadata.items()
    )
    sections = []
    for name, records in tables.items():
        sections.append(
            f"<h2>{html.escape(name.replace('_', ' ').title())}</h2>"
            "<div class='table-wrap'>"
            f"{html_table(records=list(records), max_rows=max_rows)}"
            "</div>"
        )
    path.write_text(
        f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{html.escape(title)}</title>
<style>{css}</style>
</head>
<body>
<h1>{html.escape(title)}</h1>
<p class="note">
This report combines post-build benchmark interpretation outputs. It can include
LCA summaries, AI-ready training data, and optional call-calibrator evaluation.
The database build itself is intentionally separate from these post-screening steps.
</p>
{cards}
{''.join(sections)}
</body>
</html>
""",
        encoding="utf-8",
    )
    if logger:
        logger.info("Wrote HTML report: %s", path)


def write_excel_workbook(
    *,
    tables: Mapping[str, Sequence[Mapping[str, object]]],
    output_path: str | Path,
    max_rows: int = REPORT_PREVIEW_ROWS,
    logger: logging.Logger | None = None,
) -> None:
    """Write a formatted Excel workbook for benchmark post-processing.

    Parameters
    ----------
    tables : mapping
        Table name to records.
    output_path : str or pathlib.Path
        XLSX output path.
    max_rows : int, optional
        Maximum rows per sheet. This keeps huge training tables reviewable.
    logger : logging.Logger or None, optional
        Logger.

    Raises
    ------
    ImportError
        If the optional ``openpyxl`` dependency is unavailable.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover - depends on optional extra
        raise ImportError(
            "Writing Excel reports requires openpyxl. Install with the reporting extra."
        ) from exc

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for raw_name, records in tables.items():
        sheet_name = raw_name[:31] or "table"
        worksheet = workbook.create_sheet(title=sheet_name)
        rows = [dict(record) for record in records]
        if not rows:
            worksheet.append(["No rows available"])
            continue
        fieldnames = list(rows[0].keys())
        worksheet.append(fieldnames)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for record in rows[:max_rows]:
            worksheet.append([record.get(field, "") for field in fieldnames])
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "")
            max_length = max(
                [len(header)]
                + [len(str(cell.value or "")) for cell in column_cells[1:51]]
            )
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_length + 2, 10),
                50,
            )
    workbook.save(path)
    if logger:
        logger.info("Wrote Excel workbook: %s", path)


def run_benchmark_postprocess(
    *,
    summary_dir: str | Path | None,
    out_dir: str | Path,
    calls_table: str | Path | None = None,
    taxonomy_dir: str | Path | None = None,
    taxon_map_table: str | Path | None = None,
    taxid_column: str | None = None,
    taxon_name_column: str | None = None,
    taxon_map_name_column: str | None = None,
    taxon_map_taxid_column: str | None = None,
    min_unique_kmers: int = 1,
    min_positive_sequences: int = 1,
    min_best_k: int = 0,
    max_not_detected: int | None = 50000,
    train_calibrator: bool = False,
    test_fraction: float = 0.2,
    distance_quantile: float = 0.95,
    group_columns: Sequence[str] = ("sample_id",),
    write_excel: bool = True,
    write_html: bool = True,
    logger: logging.Logger | None = None,
) -> dict[str, Path]:
    """Run LCA and AI benchmark post-processing.

    Parameters
    ----------
    summary_dir : str or pathlib.Path or None
        Existing comparable-summary directory. Used to auto-detect the long call
        table when ``calls_table`` is not provided.
    out_dir : str or pathlib.Path
        Output directory for v0.35 post-processing files.
    calls_table : str or pathlib.Path or None, optional
        Long detection/call table. If omitted, standard names are searched in
        ``summary_dir``.
    taxonomy_dir : str or pathlib.Path or None, optional
        NCBI taxonomy directory. Required for LCA generation.
    taxon_map_table : str or pathlib.Path or None, optional
        Name-to-taxid map, usually the KmerSutra genome config.
    taxid_column : str or None, optional
        Taxid column in the call/evidence table.
    taxon_name_column : str or None, optional
        Taxon-name column in the call/evidence table.
    taxon_map_name_column : str or None, optional
        Taxon-name column in the taxon map table.
    taxon_map_taxid_column : str or None, optional
        Taxid column in the taxon map table.
    min_unique_kmers : int, optional
        Minimum unique k-mers for LCA support.
    min_positive_sequences : int, optional
        Minimum positive sequences for LCA support.
    min_best_k : int, optional
        Minimum best k value for LCA support.
    max_not_detected : int or None, optional
        Cap for not-detected rows in AI training data.
    train_calibrator : bool, optional
        Whether to train/evaluate the AI call calibrator.
    test_fraction : float, optional
        Grouped test fraction for calibration.
    distance_quantile : float, optional
        Distance quantile for open-set calibration.
    group_columns : sequence of str, optional
        Group columns for train/test split.
    write_excel : bool, optional
        Whether to write a formatted XLSX report.
    write_html : bool, optional
        Whether to write an HTML report.
    logger : logging.Logger or None, optional
        Logger.

    Returns
    -------
    dict[str, pathlib.Path]
        Important output paths.
    """
    output_root = Path(out_dir).expanduser().resolve()
    output_root.mkdir(parents=True, exist_ok=True)
    resolved_calls_table = resolve_existing_table(
        explicit_path=calls_table,
        search_dir=summary_dir,
        candidate_names=DEFAULT_CALL_TABLE_NAMES,
        description="long detection/call table",
    )
    lca_table = output_root / DEFAULT_LCA_TABLE_NAME
    training_table = output_root / DEFAULT_TRAINING_TABLE_NAME
    model_json = output_root / DEFAULT_AI_MODEL_NAME
    ai_summary_table = output_root / DEFAULT_AI_SUMMARY_NAME
    ai_evaluation_table = output_root / DEFAULT_AI_EVALUATION_NAME
    html_report = output_root / f"{DEFAULT_REPORT_STEM}.html"
    excel_report = output_root / f"{DEFAULT_REPORT_STEM}.xlsx"

    if logger:
        logger.info("Post-processing calls table: %s", resolved_calls_table)
        logger.info("Post-processing output directory: %s", output_root)

    if taxonomy_dir is None:
        raise BenchmarkPostprocessError("taxonomy_dir is required for LCA reporting")
    if taxon_map_table is None:
        raise BenchmarkPostprocessError("taxon_map_table is required for LCA reporting")

    lca_records = summarise_lca_table(
        evidence_table=resolved_calls_table,
        output_table=lca_table,
        taxonomy_dir=taxonomy_dir,
        taxon_map_table=taxon_map_table,
        taxid_column=taxid_column,
        taxon_name_column=taxon_name_column,
        taxon_map_name_column=taxon_map_name_column,
        taxon_map_taxid_column=taxon_map_taxid_column,
        min_unique_kmers=min_unique_kmers,
        min_positive_sequences=min_positive_sequences,
        min_best_k=min_best_k,
        logger=logger,
    )
    training_records = write_call_training_table_from_table(
        calls_table=resolved_calls_table,
        output_table=training_table,
        max_not_detected=max_not_detected,
        lca_table=lca_table,
        logger=logger,
    )

    ai_summary_records: list[dict[str, object]] = []
    ai_evaluation_records: list[dict[str, object]] = []
    if train_calibrator:
        _model, _predictions, ai_evaluation_records = train_evaluate_call_calibrator(
            training_table=training_table,
            model_json=model_json,
            summary_table=ai_summary_table,
            evaluation_table=ai_evaluation_table,
            test_fraction=test_fraction,
            distance_quantile=distance_quantile,
            group_columns=list(group_columns),
            logger=logger,
        )
        ai_summary_records = read_records_table(input_path=ai_summary_table, logger=logger)
    else:
        if logger:
            logger.info("Skipping AI calibrator training because train_calibrator is false")

    report_tables: dict[str, Sequence[Mapping[str, object]]] = {
        "lca_scope_summary": summarise_lca_scopes(lca_records=lca_records),
        "ai_label_counts": count_records_by_column(
            records=training_records,
            column="ml_report_label",
        ),
        "ai_lca_feature_summary": summarise_numeric_features(records=training_records),
        "lca_summary_preview": lca_records[:REPORT_PREVIEW_ROWS],
        "ai_training_preview": training_records[:REPORT_PREVIEW_ROWS],
    }
    if ai_summary_records:
        report_tables["ai_training_summary"] = ai_summary_records
    if ai_evaluation_records:
        report_tables["ai_evaluation"] = ai_evaluation_records

    metadata = {
        "calls_table": resolved_calls_table,
        "lca_table": lca_table,
        "ai_training_table": training_table,
        "n_lca_rows": len(lca_records),
        "n_ai_training_rows": len(training_records),
        "train_calibrator": train_calibrator,
    }
    if write_html:
        write_html_report(
            tables=report_tables,
            output_path=html_report,
            metadata=metadata,
            logger=logger,
        )
    if write_excel:
        write_excel_workbook(
            tables=report_tables,
            output_path=excel_report,
            logger=logger,
        )

    outputs = {
        "calls_table": resolved_calls_table,
        "lca_table": lca_table,
        "ai_training_table": training_table,
        "html_report": html_report,
        "excel_report": excel_report,
    }
    if train_calibrator:
        outputs.update(
            {
                "ai_model_json": model_json,
                "ai_summary_table": ai_summary_table,
                "ai_evaluation_table": ai_evaluation_table,
            }
        )
    return outputs
