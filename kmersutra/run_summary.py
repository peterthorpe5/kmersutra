"""Run-level summary reports for KmerSutra spike-in outputs."""

from __future__ import annotations

import html
import logging
from collections import Counter, defaultdict
from pathlib import Path

from kmersutra.io import read_tsv


def _safe_int(value: object) -> int:
    """Convert a value to int, returning zero for missing values."""
    try:
        return int(float(str(value)))
    except (TypeError, ValueError):
        return 0


def _safe_float(value: object) -> float:
    """Convert a value to float, returning zero for missing values."""
    try:
        return float(str(value))
    except (TypeError, ValueError):
        return 0.0


def species_from_call_column(*, column: str) -> str:
    """Extract a species label from a KmerSutra call column name.

    Parameters
    ----------
    column : str
        Summary column such as ``kmersutra_Plasmodium_vivax_call``.

    Returns
    -------
    str
        Species label with underscores converted to spaces.
    """
    prefix = "kmersutra_"
    suffix = "_call"
    if not column.startswith(prefix) or not column.endswith(suffix):
        raise ValueError(f"Not a KmerSutra call column: {column}")
    return column[len(prefix) : -len(suffix)].replace("_", " ")


def find_species_labels(*, records: list[dict[str, str]]) -> list[str]:
    """Find species labels encoded in a KmerSutra run summary.

    Parameters
    ----------
    records : list[dict[str, str]]
        Parsed summary records.

    Returns
    -------
    list[str]
        Species labels in the order they appear in the table.
    """
    if not records:
        return []
    labels: list[str] = []
    for column in records[0]:
        if column.startswith("kmersutra_") and column.endswith("_call"):
            labels.append(species_from_call_column(column=column))
    return labels


def build_species_long_summary(*, records: list[dict[str, str]]) -> list[dict[str, object]]:
    """Convert a wide KmerSutra spike-in summary to species-long rows.

    Parameters
    ----------
    records : list[dict[str, str]]
        Parsed summary table.

    Returns
    -------
    list[dict[str, object]]
        One row per replicate, spike level and species.
    """
    species_labels = find_species_labels(records=records)
    long_rows: list[dict[str, object]] = []
    for row in records:
        for species in species_labels:
            safe_species = species.replace(" ", "_")
            long_rows.append(
                {
                    "replicate": _safe_int(row.get("replicate", 0)),
                    "spike_n": _safe_int(row.get("spike_n", 0)),
                    "total_spiked_reads": _safe_int(row.get("total_spiked_reads", 0)),
                    "species_name": species,
                    "call": row.get(f"kmersutra_{safe_species}_call", ""),
                    "n_unique_kmers": _safe_int(row.get(f"kmersutra_{safe_species}_unique_kmers", 0)),
                    "n_positive_sequences": _safe_int(row.get(f"kmersutra_{safe_species}_positive_reads", 0)),
                    "confidence_score": _safe_float(row.get(f"kmersutra_{safe_species}_confidence", 0)),
                    "kmersutra_runtime_seconds": _safe_int(row.get("kmersutra_runtime_seconds", 0)),
                    "kmersutra_out_dir": row.get("kmersutra_out_dir", ""),
                    "kmersutra_calls_tsv": row.get("kmersutra_calls_tsv", ""),
                    "kmersutra_evidence_tsv": row.get("kmersutra_evidence_tsv", ""),
                    "kmersutra_read_hits_tsv_gz": row.get("kmersutra_read_hits_tsv_gz", ""),
                }
            )
    return long_rows


def summarise_call_counts(*, species_long_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    """Summarise call counts by species and call class.

    Parameters
    ----------
    species_long_rows : list[dict[str, object]]
        Long-format species summary records.

    Returns
    -------
    list[dict[str, object]]
        Call-count summary rows.
    """
    counter: Counter[tuple[str, str]] = Counter()
    for row in species_long_rows:
        counter[(str(row["species_name"]), str(row["call"]))] += 1
    return [
        {"species_name": species, "call": call, "n_records": count}
        for (species, call), count in sorted(counter.items())
    ]


def summarise_by_spike(*, species_long_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    """Summarise calls and evidence by spike level and species.

    Parameters
    ----------
    species_long_rows : list[dict[str, object]]
        Long-format species summary records.

    Returns
    -------
    list[dict[str, object]]
        Spike/species summary rows.
    """
    grouped: dict[tuple[int, str], list[dict[str, object]]] = defaultdict(list)
    for row in species_long_rows:
        grouped[(_safe_int(row["spike_n"]), str(row["species_name"]))].append(row)

    out_rows: list[dict[str, object]] = []
    for (spike_n, species_name), rows in sorted(grouped.items()):
        call_counts = Counter(str(row["call"]) for row in rows)
        present_count = sum(
            count
            for call, count in call_counts.items()
            if call in {"present_high_confidence", "present_in_mixed_sample"}
        )
        out_rows.append(
            {
                "spike_n": spike_n,
                "species_name": species_name,
                "n_records": len(rows),
                "n_present_calls": present_count,
                "present_call_rate": round(present_count / len(rows), 4) if rows else 0,
                "mean_unique_kmers": round(
                    sum(_safe_int(row["n_unique_kmers"]) for row in rows) / len(rows), 4
                )
                if rows
                else 0,
                "mean_positive_sequences": round(
                    sum(_safe_int(row["n_positive_sequences"]) for row in rows) / len(rows), 4
                )
                if rows
                else 0,
                "mean_confidence_score": round(
                    sum(_safe_float(row["confidence_score"]) for row in rows) / len(rows), 4
                )
                if rows
                else 0,
                "call_counts": "; ".join(f"{call}={count}" for call, count in sorted(call_counts.items())),
            }
        )
    return out_rows


def write_run_summary_html(
    *,
    summary_records: list[dict[str, str]],
    species_long_rows: list[dict[str, object]],
    call_count_rows: list[dict[str, object]],
    by_spike_rows: list[dict[str, object]],
    output_path: str | Path,
) -> None:
    """Write an HTML run-level summary report.

    Parameters
    ----------
    summary_records : list[dict[str, str]]
        Parsed wide summary records.
    species_long_rows : list[dict[str, object]]
        Long-format species records.
    call_count_rows : list[dict[str, object]]
        Call count records.
    by_spike_rows : list[dict[str, object]]
        Spike/species summary records.
    output_path : str or pathlib.Path
        Output HTML path.
    """
    def table(title: str, records: list[dict[str, object]], max_rows: int = 500) -> str:
        if not records:
            return f"<h2>{html.escape(title)}</h2><p>No records available.</p>"
        columns = list(records[0].keys())
        head = "".join(f"<th>{html.escape(column)}</th>" for column in columns)
        rows = []
        for record in records[:max_rows]:
            cells = "".join(html.escape(str(record.get(column, ""))) for column in columns)
            cells = "".join(f"<td>{html.escape(str(record.get(column, '')))}</td>" for column in columns)
            rows.append(f"<tr>{cells}</tr>")
        return (
            f"<h2>{html.escape(title)}</h2>"
            "<div class='table-wrap'><table>"
            f"<thead><tr>{head}</tr></thead><tbody>{''.join(rows)}</tbody></table></div>"
        )

    style = """
    <style>
      body { font-family: Arial, Helvetica, sans-serif; margin: 0; background: #f7f9fc; color: #1a1a1a; }
      .container { max-width: 1500px; margin: 0 auto; padding: 32px; background: white; }
      h1, h2 { color: #1f4e79; }
      .note { background: #eef5fb; border-left: 5px solid #1f4e79; padding: 12px 16px; margin: 18px 0; }
      .table-wrap { overflow-x: auto; border: 1px solid #dbe5f0; border-radius: 8px; margin-bottom: 28px; }
      table { border-collapse: collapse; width: 100%; font-size: 13px; }
      th { background: #1f4e79; color: white; padding: 8px; text-align: left; }
      td { border-top: 1px solid #e4edf5; padding: 7px; vertical-align: top; }
      tr:nth-child(even) { background: #f9fbfd; }
    </style>
    """
    page = (
        "<!DOCTYPE html><html><head><meta charset='utf-8'>"
        f"<title>KmerSutra run summary</title>{style}</head><body><div class='container'>"
        "<h1>KmerSutra run summary</h1>"
        "<div class='note'>Calls labelled <strong>present_in_mixed_sample</strong> indicate "
        "that more than one species independently passed the evidence thresholds. This is "
        "expected in deliberately mixed spike-in samples and should not be interpreted as an "
        "off-target conflict by itself.</div>"
        + table("Call counts", call_count_rows)
        + table("By spike level", by_spike_rows)
        + table("Species-long summary", species_long_rows)
        + table("Raw wide summary", [dict(row) for row in summary_records])
        + "</div></body></html>"
    )
    Path(output_path).write_text(page, encoding="utf-8")


def write_run_summary_excel(
    *,
    summary_records: list[dict[str, str]],
    species_long_rows: list[dict[str, object]],
    call_count_rows: list[dict[str, object]],
    by_spike_rows: list[dict[str, object]],
    output_path: str | Path,
) -> None:
    """Write a formatted Excel summary workbook.

    Parameters
    ----------
    summary_records : list[dict[str, str]]
        Parsed wide summary records.
    species_long_rows : list[dict[str, object]]
        Long-format species records.
    call_count_rows : list[dict[str, object]]
        Call count records.
    by_spike_rows : list[dict[str, object]]
        Spike/species summary records.
    output_path : str or pathlib.Path
        Output XLSX path.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise ImportError("openpyxl is required to write Excel summaries") from exc

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    datasets: list[tuple[str, list[dict[str, object]]]] = [
        ("Run_Summary", [dict(row) for row in summary_records]),
        ("Species_Long", species_long_rows),
        ("By_Spike", by_spike_rows),
        ("Call_Counts", call_count_rows),
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for sheet_name, records in datasets:
        sheet = workbook.create_sheet(title=sheet_name)
        if records:
            columns = list(records[0].keys())
        else:
            columns = ["message"]
            records = [{"message": "No records available"}]
        sheet.append(columns)
        for record in records:
            sheet.append([record.get(column, "") for column in columns])

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table_name = f"{sheet_name.replace('_', '')}Table"[:240]
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap_alignment

        for col_index, column in enumerate(columns, start=1):
            values = [str(column)] + [str(record.get(column, "")) for record in records[:200]]
            max_len = max(len(value) for value in values)
            width = min(max(max_len + 2, 10), 48)
            sheet.column_dimensions[get_column_letter(col_index)].width = width

    workbook.save(output_path)


def build_run_summary_reports(
    *,
    summary_tsv: str | Path,
    out_xlsx: str | Path | None = None,
    out_html: str | Path | None = None,
    logger: logging.Logger | None = None,
) -> dict[str, Path]:
    """Build run-level Excel and HTML summaries from a KmerSutra TSV.

    Parameters
    ----------
    summary_tsv : str or pathlib.Path
        KmerSutra spike-in summary TSV.
    out_xlsx : str or pathlib.Path or None, optional
        Optional Excel output path.
    out_html : str or pathlib.Path or None, optional
        Optional HTML output path.
    logger : logging.Logger or None, optional
        Logger for progress messages.

    Returns
    -------
    dict[str, pathlib.Path]
        Paths written, keyed by output type.
    """
    summary_path = Path(summary_tsv)
    summary_records = read_tsv(input_path=summary_path)
    if logger:
        logger.info("Loaded %d KmerSutra summary rows from %s", len(summary_records), summary_path)
    species_long_rows = build_species_long_summary(records=summary_records)
    call_count_rows = summarise_call_counts(species_long_rows=species_long_rows)
    by_spike_rows = summarise_by_spike(species_long_rows=species_long_rows)

    written: dict[str, Path] = {}
    if out_xlsx is not None:
        xlsx_path = Path(out_xlsx)
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        write_run_summary_excel(
            summary_records=summary_records,
            species_long_rows=species_long_rows,
            call_count_rows=call_count_rows,
            by_spike_rows=by_spike_rows,
            output_path=xlsx_path,
        )
        written["xlsx"] = xlsx_path
        if logger:
            logger.info("Wrote Excel summary: %s", xlsx_path)
    if out_html is not None:
        html_path = Path(out_html)
        html_path.parent.mkdir(parents=True, exist_ok=True)
        write_run_summary_html(
            summary_records=summary_records,
            species_long_rows=species_long_rows,
            call_count_rows=call_count_rows,
            by_spike_rows=by_spike_rows,
            output_path=html_path,
        )
        written["html"] = html_path
        if logger:
            logger.info("Wrote HTML summary: %s", html_path)
    return written
