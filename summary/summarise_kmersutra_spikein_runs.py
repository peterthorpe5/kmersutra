#!/usr/bin/env python3
"""
Summarise KmerSutra spike-in benchmark runs.

This script collects one or more KmerSutra spike-in run directories and writes
combined TSV, formatted Excel, and HTML summaries. It is designed for outputs
from the KmerSutra spike-in shell workflow, where each run contains a main
`spikein_multi_kmersutra_summary.tsv` file and each replicate/spike directory
contains KmerSutra per-sample output files.

The script intentionally writes tab-separated outputs rather than CSV files.
"""

from __future__ import annotations

import argparse
import html
import logging
import math
import re
import sys
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

LOGGER = logging.getLogger("summarise_kmersutra_spikein_runs")

SUMMARY_FILE_NAME = "spikein_multi_kmersutra_summary.tsv"
SPECIES_COLUMN_PATTERN = re.compile(
    r"^kmersutra_(?P<label>.+)_(?P<metric>call|unique_kmers|positive_reads|confidence)$"
)
POSITIVE_CALLS = {
    "present_high_confidence",
    "present_low_confidence",
    "present_in_mixed_sample",
    "mixed_species_present",
}
AMBIGUOUS_CALLS = {
    "ambiguous_mixed_signal",
    "ambiguous_conflicting_signal",
}
NOT_DETECTED_CALLS = {"not_detected", "NA", "", "nan"}


def parse_args() -> argparse.Namespace:
    """Parse command-line arguments.

    Returns
    -------
    argparse.Namespace
        Parsed command-line arguments.
    """
    parser = argparse.ArgumentParser(
        description=(
            "Summarise KmerSutra spike-in run folders into combined TSV, "
            "formatted Excel, and HTML outputs."
        )
    )
    parser.add_argument(
        "--input_dirs",
        nargs="+",
        required=True,
        help=(
            "One or more KmerSutra run directories, or parent directories "
            "containing KmerSutra run directories."
        ),
    )
    parser.add_argument(
        "--out_dir",
        required=True,
        help="Output directory for combined summaries.",
    )
    parser.add_argument(
        "--summary_name",
        default=SUMMARY_FILE_NAME,
        help=f"Main run summary filename to search for. Default: {SUMMARY_FILE_NAME}",
    )
    parser.add_argument(
        "--run_glob",
        default="spikein_multi_kmersutra*",
        help=(
            "Directory-name pattern used when searching parent directories. "
            "Default: spikein_multi_kmersutra*"
        ),
    )
    parser.add_argument(
        "--expected_replicates",
        type=int,
        default=None,
        help="Optional expected number of replicates for run QC.",
    )
    parser.add_argument(
        "--expected_spike_levels",
        default=None,
        help=(
            "Optional space-separated expected spike levels, for example "
            "'0 1 5 10 25 50 100 250 500 1000 2500 5000'."
        ),
    )
    parser.add_argument(
        "--max_html_rows",
        type=int,
        default=500,
        help="Maximum rows per table shown in the HTML report.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print more detailed logging.",
    )
    return parser.parse_args()


def configure_logging(*, verbose: bool) -> None:
    """Configure logging for command-line execution.

    Parameters
    ----------
    verbose : bool
        Whether to use INFO-level logging.
    """
    level = logging.INFO if verbose else logging.WARNING
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def read_tsv_file(*, path: Path) -> pd.DataFrame:
    """Read a tab-separated file into a dataframe.

    Parameters
    ----------
    path : Path
        Path to the TSV file.

    Returns
    -------
    pd.DataFrame
        Loaded dataframe. Empty files are returned as empty dataframes.
    """
    if not path.exists() or path.stat().st_size == 0:
        LOGGER.warning("Missing or empty TSV: %s", path)
        return pd.DataFrame()
    try:
        return pd.read_csv(path, sep="\t", dtype=str)
    except pd.errors.EmptyDataError:
        LOGGER.warning("Empty TSV: %s", path)
        return pd.DataFrame()


def write_tsv_file(*, dataframe: pd.DataFrame, path: Path) -> None:
    """Write a dataframe as a tab-separated file.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Dataframe to write.
    path : Path
        Output TSV path.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(path, sep="\t", index=False)
    LOGGER.info("Wrote TSV: %s", path)


def parse_expected_spike_levels(*, raw_value: str | None) -> list[float] | None:
    """Parse optional expected spike levels.

    Parameters
    ----------
    raw_value : str | None
        Space-separated spike levels, or None.

    Returns
    -------
    list[float] | None
        Parsed spike levels, or None if not provided.
    """
    if raw_value is None:
        return None
    values = []
    for token in raw_value.split():
        try:
            values.append(float(token))
        except ValueError as exc:
            raise ValueError(f"Invalid spike level: {token}") from exc
    return values


def normalise_numeric_columns(
    *,
    dataframe: pd.DataFrame,
    columns: Iterable[str],
) -> pd.DataFrame:
    """Convert selected dataframe columns to numeric values.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Input dataframe.
    columns : Iterable[str]
        Column names to convert if present.

    Returns
    -------
    pd.DataFrame
        Dataframe with converted columns.
    """
    out_df = dataframe.copy()
    for column in columns:
        if column in out_df.columns:
            out_df[column] = pd.to_numeric(out_df[column], errors="coerce")
    return out_df


def find_summary_files(
    *,
    input_dirs: list[str],
    summary_name: str,
    run_glob: str,
) -> list[Path]:
    """Find KmerSutra run summary files.

    Parameters
    ----------
    input_dirs : list[str]
        Run directories or parent directories.
    summary_name : str
        Summary filename to find.
    run_glob : str
        Directory-name pattern for KmerSutra run directories.

    Returns
    -------
    list[Path]
        Sorted unique summary file paths.
    """
    summary_paths: set[Path] = set()
    for raw_dir in input_dirs:
        input_path = Path(raw_dir).expanduser().resolve()
        if not input_path.exists():
            LOGGER.warning("Input path does not exist: %s", input_path)
            continue

        direct_summary = input_path / summary_name
        if direct_summary.exists():
            summary_paths.add(direct_summary)
            continue

        for run_dir in input_path.glob(run_glob):
            candidate = run_dir / summary_name
            if candidate.exists():
                summary_paths.add(candidate.resolve())

        if not summary_paths:
            for candidate in input_path.rglob(summary_name):
                summary_paths.add(candidate.resolve())

    sorted_paths = sorted(summary_paths)
    LOGGER.info("Found %d summary files", len(sorted_paths))
    for summary_path in sorted_paths:
        LOGGER.info("Summary file: %s", summary_path)
    return sorted_paths


def load_run_summaries(*, summary_paths: list[Path]) -> pd.DataFrame:
    """Load and combine KmerSutra run summary tables.

    Parameters
    ----------
    summary_paths : list[Path]
        Paths to run-level summary TSV files.

    Returns
    -------
    pd.DataFrame
        Combined run summary dataframe.
    """
    frames = []
    for summary_path in summary_paths:
        run_df = read_tsv_file(path=summary_path)
        if run_df.empty:
            continue
        run_dir = summary_path.parent.resolve()
        run_df.insert(0, "run_name", run_dir.name)
        run_df.insert(1, "run_dir", str(run_dir))
        run_df.insert(2, "summary_tsv", str(summary_path.resolve()))
        frames.append(run_df)
    if not frames:
        return pd.DataFrame()
    combined = pd.concat(frames, ignore_index=True)
    combined = normalise_numeric_columns(
        dataframe=combined,
        columns=[
            "replicate",
            "spike_n",
            "total_spiked_reads",
            "kmersutra_runtime_seconds",
        ],
    )
    return combined


def safe_species_name_from_label(*, label: str) -> str:
    """Convert a shell-safe species label back to a readable label.

    Parameters
    ----------
    label : str
        Shell-safe label, for example `Plasmodium_vivax`.

    Returns
    -------
    str
        Readable label, for example `Plasmodium vivax`.
    """
    return label.replace("_", " ").strip()


def extract_species_labels_from_summary(*, summary_df: pd.DataFrame) -> list[str]:
    """Extract species labels from wide KmerSutra summary columns.

    Parameters
    ----------
    summary_df : pd.DataFrame
        Combined run summary dataframe.

    Returns
    -------
    list[str]
        Sorted species labels.
    """
    labels = set()
    for column in summary_df.columns:
        match = SPECIES_COLUMN_PATTERN.match(column)
        if match:
            labels.add(match.group("label"))
    return sorted(labels)


def build_species_long_from_wide(*, summary_df: pd.DataFrame) -> pd.DataFrame:
    """Build a long per-species table from the wide run summary.

    Parameters
    ----------
    summary_df : pd.DataFrame
        Combined wide run summary dataframe.

    Returns
    -------
    pd.DataFrame
        Long per-species summary table.
    """
    if summary_df.empty:
        return pd.DataFrame()

    rows = []
    species_labels = extract_species_labels_from_summary(summary_df=summary_df)
    LOGGER.info("Extracted %d species labels from wide summary", len(species_labels))

    for _, row in summary_df.iterrows():
        for safe_label in species_labels:
            prefix = f"kmersutra_{safe_label}_"
            rows.append(
                {
                    "run_name": row.get("run_name"),
                    "run_dir": row.get("run_dir"),
                    "replicate": row.get("replicate"),
                    "spike_n": row.get("spike_n"),
                    "total_spiked_reads": row.get("total_spiked_reads"),
                    "sample_id": f"rep{row.get('replicate')}_n{row.get('spike_n')}",
                    "species_name": safe_species_name_from_label(label=safe_label),
                    "call": row.get(prefix + "call", "NA"),
                    "n_unique_kmers": row.get(prefix + "unique_kmers", 0),
                    "n_positive_sequences": row.get(prefix + "positive_reads", 0),
                    "confidence_score": row.get(prefix + "confidence", 0),
                    "kmersutra_out_dir": row.get("kmersutra_out_dir"),
                    "kmersutra_calls_tsv": row.get("kmersutra_calls_tsv"),
                    "kmersutra_runtime_seconds": row.get("kmersutra_runtime_seconds"),
                }
            )

    out_df = pd.DataFrame(rows)
    out_df = normalise_numeric_columns(
        dataframe=out_df,
        columns=[
            "replicate",
            "spike_n",
            "total_spiked_reads",
            "n_unique_kmers",
            "n_positive_sequences",
            "confidence_score",
            "kmersutra_runtime_seconds",
        ],
    )
    return out_df


def resolve_output_path(*, raw_path: object, run_dir: Path) -> Path | None:
    """Resolve an output path recorded in the run summary.

    Parameters
    ----------
    raw_path : object
        Raw path value from a summary dataframe cell.
    run_dir : Path
        Run directory used as a fallback base.

    Returns
    -------
    Path | None
        Existing path if found, otherwise None.
    """
    if raw_path is None or pd.isna(raw_path):
        return None
    raw_text = str(raw_path).strip()
    if not raw_text or raw_text.upper() == "NA":
        return None

    candidate = Path(raw_text)
    if candidate.exists():
        return candidate

    # Fallback: search by basename below the run directory. This helps when
    # paths were copied back from a job-local directory but not rewritten.
    matches = list(run_dir.rglob(candidate.name))
    if matches:
        return matches[0]
    LOGGER.debug("Could not resolve path: %s", raw_text)
    return None


def load_detection_call_files(*, summary_df: pd.DataFrame) -> pd.DataFrame:
    """Load per-sample KmerSutra detection call files.

    Parameters
    ----------
    summary_df : pd.DataFrame
        Combined run summary dataframe.

    Returns
    -------
    pd.DataFrame
        Combined detection call table.
    """
    if summary_df.empty or "kmersutra_calls_tsv" not in summary_df.columns:
        return pd.DataFrame()

    frames = []
    seen_paths: set[Path] = set()
    for _, row in summary_df.iterrows():
        run_dir = Path(str(row.get("run_dir", "."))).resolve()
        call_path = resolve_output_path(
            raw_path=row.get("kmersutra_calls_tsv"),
            run_dir=run_dir,
        )
        if call_path is None or call_path in seen_paths:
            continue
        seen_paths.add(call_path)

        call_df = read_tsv_file(path=call_path)
        if call_df.empty:
            continue
        call_df.insert(0, "run_name", row.get("run_name"))
        call_df.insert(1, "run_dir", row.get("run_dir"))
        call_df.insert(2, "replicate", row.get("replicate"))
        call_df.insert(3, "spike_n", row.get("spike_n"))
        call_df.insert(4, "total_spiked_reads", row.get("total_spiked_reads"))
        call_df.insert(5, "calls_tsv", str(call_path.resolve()))
        frames.append(call_df)

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)
    combined = normalise_numeric_columns(
        dataframe=combined,
        columns=[
            "replicate",
            "spike_n",
            "total_spiked_reads",
            "n_hits",
            "n_unique_kmers",
            "n_positive_sequences",
            "n_k_values_positive",
            "best_k",
            "n_exact_hits",
            "n_fuzzy_hits",
            "conflicting_unique_kmers",
            "conflict_ratio",
            "confidence_score",
        ],
    )
    return combined


def load_hit_summary_files(*, summary_df: pd.DataFrame) -> pd.DataFrame:
    """Load per-sample KmerSutra hit summary files.

    Parameters
    ----------
    summary_df : pd.DataFrame
        Combined run summary dataframe.

    Returns
    -------
    pd.DataFrame
        Combined hit summary dataframe.
    """
    if summary_df.empty or "kmersutra_out_dir" not in summary_df.columns:
        return pd.DataFrame()

    frames = []
    seen_paths: set[Path] = set()
    for _, row in summary_df.iterrows():
        run_dir = Path(str(row.get("run_dir", "."))).resolve()
        out_dir = resolve_output_path(raw_path=row.get("kmersutra_out_dir"), run_dir=run_dir)
        if out_dir is None:
            continue
        hit_path = out_dir / "sample_species_kmer_hits.tsv"
        if not hit_path.exists():
            hit_path = out_dir / "sample_species_kmer_evidence.tsv"
        if not hit_path.exists() or hit_path in seen_paths:
            continue
        seen_paths.add(hit_path)

        hit_df = read_tsv_file(path=hit_path)
        if hit_df.empty:
            continue
        hit_df.insert(0, "run_name", row.get("run_name"))
        hit_df.insert(1, "run_dir", row.get("run_dir"))
        hit_df.insert(2, "replicate", row.get("replicate"))
        hit_df.insert(3, "spike_n", row.get("spike_n"))
        hit_df.insert(4, "total_spiked_reads", row.get("total_spiked_reads"))
        hit_df.insert(5, "source_tsv", str(hit_path.resolve()))
        frames.append(hit_df)

    if not frames:
        return pd.DataFrame()
    combined = pd.concat(frames, ignore_index=True)
    numeric_cols = [column for column in combined.columns if column.startswith("n_")]
    numeric_cols += ["replicate", "spike_n", "total_spiked_reads", "k"]
    return normalise_numeric_columns(dataframe=combined, columns=numeric_cols)


def classify_call_state(*, call_value: object) -> str:
    """Classify a KmerSutra call into broad states.

    Parameters
    ----------
    call_value : object
        Raw call value.

    Returns
    -------
    str
        Broad call state: positive, ambiguous, not_detected, or other.
    """
    call_text = str(call_value).strip()
    if call_text in POSITIVE_CALLS:
        return "positive"
    if call_text in AMBIGUOUS_CALLS:
        return "ambiguous"
    if call_text in NOT_DETECTED_CALLS:
        return "not_detected"
    return "other"


def choose_authoritative_species_table(
    *,
    species_long_df: pd.DataFrame,
    detection_calls_df: pd.DataFrame,
) -> pd.DataFrame:
    """Choose the best available per-species table for summaries.

    Parameters
    ----------
    species_long_df : pd.DataFrame
        Long table derived from the wide shell summary.
    detection_calls_df : pd.DataFrame
        Combined per-sample detection call files.

    Returns
    -------
    pd.DataFrame
        Authoritative per-species table.
    """
    if not detection_calls_df.empty:
        out_df = detection_calls_df.copy()
        # Normalise naming in case older outputs used n_positive_reads.
        if "n_positive_sequences" not in out_df.columns and "n_positive_reads" in out_df.columns:
            out_df["n_positive_sequences"] = out_df["n_positive_reads"]
        return out_df
    return species_long_df.copy()


def build_by_spike_summary(*, species_df: pd.DataFrame) -> pd.DataFrame:
    """Summarise KmerSutra calls by species and spike level.

    Parameters
    ----------
    species_df : pd.DataFrame
        Authoritative per-species table.

    Returns
    -------
    pd.DataFrame
        By-spike species summary.
    """
    if species_df.empty:
        return pd.DataFrame()

    work_df = species_df.copy()
    work_df["call_state"] = work_df["call"].apply(classify_call_state)
    work_df["is_positive_call"] = work_df["call_state"].eq("positive")
    work_df["is_ambiguous_call"] = work_df["call_state"].eq("ambiguous")
    work_df = normalise_numeric_columns(
        dataframe=work_df,
        columns=[
            "spike_n",
            "n_unique_kmers",
            "n_positive_sequences",
            "confidence_score",
            "conflict_ratio",
            "kmersutra_runtime_seconds",
        ],
    )

    group_columns = ["species_name", "spike_n"]
    if "clade" in work_df.columns:
        group_columns.insert(1, "clade")

    summary_df = (
        work_df.groupby(group_columns, dropna=False)
        .agg(
            n_observations=("call", "size"),
            n_positive_calls=("is_positive_call", "sum"),
            n_ambiguous_calls=("is_ambiguous_call", "sum"),
            mean_unique_kmers=("n_unique_kmers", "mean"),
            median_unique_kmers=("n_unique_kmers", "median"),
            mean_positive_sequences=("n_positive_sequences", "mean"),
            median_positive_sequences=("n_positive_sequences", "median"),
            mean_confidence_score=("confidence_score", "mean"),
            median_confidence_score=("confidence_score", "median"),
            mean_conflict_ratio=("conflict_ratio", "mean")
            if "conflict_ratio" in work_df.columns
            else ("confidence_score", "size"),
        )
        .reset_index()
    )
    summary_df["positive_call_rate"] = (
        summary_df["n_positive_calls"] / summary_df["n_observations"]
    )
    summary_df["ambiguous_call_rate"] = (
        summary_df["n_ambiguous_calls"] / summary_df["n_observations"]
    )
    return summary_df.sort_values(group_columns).reset_index(drop=True)


def build_call_counts(*, species_df: pd.DataFrame) -> pd.DataFrame:
    """Build call-count summaries by species, spike level, and call.

    Parameters
    ----------
    species_df : pd.DataFrame
        Authoritative per-species table.

    Returns
    -------
    pd.DataFrame
        Call count table.
    """
    if species_df.empty:
        return pd.DataFrame()
    group_cols = ["species_name", "spike_n", "call"]
    if "clade" in species_df.columns:
        group_cols.insert(1, "clade")
    counts_df = species_df.groupby(group_cols, dropna=False).size().reset_index(name="n")
    return counts_df.sort_values(group_cols).reset_index(drop=True)


def build_run_qc(
    *,
    summary_df: pd.DataFrame,
    expected_replicates: int | None,
    expected_spike_levels: list[float] | None,
) -> pd.DataFrame:
    """Build a run-level quality-control table.

    Parameters
    ----------
    summary_df : pd.DataFrame
        Combined run summary dataframe.
    expected_replicates : int | None
        Optional expected replicate count.
    expected_spike_levels : list[float] | None
        Optional expected spike levels.

    Returns
    -------
    pd.DataFrame
        Run-level QC table.
    """
    if summary_df.empty:
        return pd.DataFrame()

    rows = []
    for run_name, run_df in summary_df.groupby("run_name", dropna=False):
        spike_values = sorted(pd.to_numeric(run_df["spike_n"], errors="coerce").dropna().unique())
        replicate_values = sorted(pd.to_numeric(run_df["replicate"], errors="coerce").dropna().unique())
        n_expected_rows = math.nan
        if expected_replicates is not None and expected_spike_levels is not None:
            n_expected_rows = expected_replicates * len(expected_spike_levels)
        rows.append(
            {
                "run_name": run_name,
                "run_dir": run_df["run_dir"].iloc[0] if "run_dir" in run_df.columns else "",
                "n_rows": int(run_df.shape[0]),
                "n_expected_rows": n_expected_rows,
                "n_replicates_observed": len(replicate_values),
                "n_spike_levels_observed": len(spike_values),
                "spike_levels_observed": " ".join(str(int(x)) if x.is_integer() else str(x) for x in spike_values),
                "expected_replicates": expected_replicates,
                "expected_spike_levels": " ".join(str(x) for x in expected_spike_levels)
                if expected_spike_levels is not None
                else "",
                "missing_expected_rows": (
                    int(n_expected_rows - run_df.shape[0])
                    if not pd.isna(n_expected_rows)
                    else math.nan
                ),
            }
        )
    return pd.DataFrame(rows)


def dataframe_to_html_table(*, dataframe: pd.DataFrame, max_rows: int) -> str:
    """Render a dataframe as an HTML table.

    Parameters
    ----------
    dataframe : pd.DataFrame
        Dataframe to render.
    max_rows : int
        Maximum number of rows to render.

    Returns
    -------
    str
        HTML table or placeholder paragraph.
    """
    if dataframe.empty:
        return "<p>No records available.</p>"
    display_df = dataframe.head(max_rows).copy()
    table_html = display_df.to_html(index=False, escape=True, border=0)
    note = ""
    if dataframe.shape[0] > max_rows:
        note = f"<p><em>Showing first {max_rows} of {dataframe.shape[0]} rows.</em></p>"
    return f"{note}<div class='table-wrap'>{table_html}</div>"


def build_html_report(
    *,
    out_path: Path,
    run_qc_df: pd.DataFrame,
    by_spike_df: pd.DataFrame,
    call_counts_df: pd.DataFrame,
    species_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    max_html_rows: int,
) -> None:
    """Write an HTML summary report.

    Parameters
    ----------
    out_path : Path
        Output HTML path.
    run_qc_df : pd.DataFrame
        Run QC table.
    by_spike_df : pd.DataFrame
        By-spike species summary.
    call_counts_df : pd.DataFrame
        Call-count table.
    species_df : pd.DataFrame
        Per-species long table.
    summary_df : pd.DataFrame
        Combined wide run summary.
    max_html_rows : int
        Maximum rows to show for long tables.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    n_runs = summary_df["run_name"].nunique() if "run_name" in summary_df.columns else 0
    n_samples = summary_df.shape[0]
    n_species_records = species_df.shape[0]

    html_text = f"""<!DOCTYPE html>
<html>
<head>
<meta charset=\"utf-8\">
<title>KmerSutra spike-in summary</title>
<style>
body {{ font-family: Arial, Helvetica, sans-serif; margin: 0; background: #f7f9fc; color: #1a1a1a; }}
.container {{ max-width: 1500px; margin: 0 auto; padding: 32px; background: #ffffff; }}
h1, h2, h3 {{ color: #1f4e79; }}
.note {{ background: #eef5fb; border-left: 5px solid #1f4e79; padding: 12px 16px; margin: 18px 0; }}
.metric-grid {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 14px; margin: 18px 0; }}
.metric-card {{ background: #f7f9fc; border: 1px solid #dbe5f0; border-radius: 8px; padding: 14px; }}
.metric-value {{ font-size: 24px; font-weight: bold; color: #1f4e79; }}
.table-wrap {{ overflow-x: auto; border: 1px solid #dbe5f0; border-radius: 8px; margin-bottom: 28px; }}
table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
th {{ background: #1f4e79; color: white; padding: 8px; text-align: left; position: sticky; top: 0; }}
td {{ border-top: 1px solid #e4edf5; padding: 7px; vertical-align: top; }}
tr:nth-child(even) {{ background: #f9fbfd; }}
</style>
</head>
<body>
<div class=\"container\">
<h1>KmerSutra spike-in summary</h1>
<div class=\"note\">
This report summarises KmerSutra spike-in runs. Confidence scores are heuristic unless calibrated against held-out spike-in truth data.
</div>
<div class=\"metric-grid\">
  <div class=\"metric-card\"><div>Runs</div><div class=\"metric-value\">{n_runs}</div></div>
  <div class=\"metric-card\"><div>Samples</div><div class=\"metric-value\">{n_samples}</div></div>
  <div class=\"metric-card\"><div>Species-level records</div><div class=\"metric-value\">{n_species_records}</div></div>
</div>
<h2>Run QC</h2>
{dataframe_to_html_table(dataframe=run_qc_df, max_rows=max_html_rows)}
<h2>By-spike species summary</h2>
{dataframe_to_html_table(dataframe=by_spike_df, max_rows=max_html_rows)}
<h2>Call counts</h2>
{dataframe_to_html_table(dataframe=call_counts_df, max_rows=max_html_rows)}
<h2>Per-species records</h2>
{dataframe_to_html_table(dataframe=species_df, max_rows=max_html_rows)}
<h2>Combined run summary</h2>
{dataframe_to_html_table(dataframe=summary_df, max_rows=max_html_rows)}
</div>
</body>
</html>
"""
    out_path.write_text(html_text, encoding="utf-8")
    LOGGER.info("Wrote HTML report: %s", out_path)


def write_formatted_excel(
    *,
    out_path: Path,
    sheet_dataframes: dict[str, pd.DataFrame],
) -> None:
    """Write a formatted Excel workbook.

    Parameters
    ----------
    out_path : Path
        Output Excel path.
    sheet_dataframes : dict[str, pd.DataFrame]
        Mapping from sheet names to dataframes.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, dataframe in sheet_dataframes.items():
            clean_name = sheet_name[:31]
            export_df = dataframe.copy()
            if export_df.empty:
                export_df = pd.DataFrame({"message": ["No records available"]})
            export_df.to_excel(writer, sheet_name=clean_name, index=False)

    workbook = load_workbook(out_path)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        if max_row >= 1 and max_col >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 60)
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=False, vertical="top")
    workbook.save(out_path)
    LOGGER.info("Wrote formatted Excel workbook: %s", out_path)


def run_summary_workflow(
    *,
    input_dirs: list[str],
    out_dir: Path,
    summary_name: str,
    run_glob: str,
    expected_replicates: int | None,
    expected_spike_levels: list[float] | None,
    max_html_rows: int,
) -> None:
    """Run the KmerSutra spike-in summary workflow.

    Parameters
    ----------
    input_dirs : list[str]
        Input run or parent directories.
    out_dir : Path
        Output directory.
    summary_name : str
        Run summary filename to search for.
    run_glob : str
        Run directory glob used when searching parent directories.
    expected_replicates : int | None
        Optional expected replicate count.
    expected_spike_levels : list[float] | None
        Optional expected spike levels.
    max_html_rows : int
        Maximum rows to show in HTML tables.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    summary_paths = find_summary_files(
        input_dirs=input_dirs,
        summary_name=summary_name,
        run_glob=run_glob,
    )
    if not summary_paths:
        raise FileNotFoundError("No KmerSutra run summary files were found.")

    summary_df = load_run_summaries(summary_paths=summary_paths)
    if summary_df.empty:
        raise ValueError("No non-empty KmerSutra summary files were loaded.")

    species_long_wide_df = build_species_long_from_wide(summary_df=summary_df)
    detection_calls_df = load_detection_call_files(summary_df=summary_df)
    hit_summary_df = load_hit_summary_files(summary_df=summary_df)
    species_df = choose_authoritative_species_table(
        species_long_df=species_long_wide_df,
        detection_calls_df=detection_calls_df,
    )
    by_spike_df = build_by_spike_summary(species_df=species_df)
    call_counts_df = build_call_counts(species_df=species_df)
    run_qc_df = build_run_qc(
        summary_df=summary_df,
        expected_replicates=expected_replicates,
        expected_spike_levels=expected_spike_levels,
    )

    write_tsv_file(dataframe=summary_df, path=out_dir / "combined_run_summary.tsv")
    write_tsv_file(dataframe=species_long_wide_df, path=out_dir / "species_long_from_wide_summary.tsv")
    write_tsv_file(dataframe=detection_calls_df, path=out_dir / "combined_detection_calls.tsv")
    write_tsv_file(dataframe=hit_summary_df, path=out_dir / "combined_hit_summary.tsv")
    write_tsv_file(dataframe=species_df, path=out_dir / "authoritative_species_summary.tsv")
    write_tsv_file(dataframe=by_spike_df, path=out_dir / "by_spike_species_summary.tsv")
    write_tsv_file(dataframe=call_counts_df, path=out_dir / "call_counts.tsv")
    write_tsv_file(dataframe=run_qc_df, path=out_dir / "run_qc.tsv")

    write_formatted_excel(
        out_path=out_dir / "kmersutra_spikein_overall_summary.xlsx",
        sheet_dataframes={
            "Run_QC": run_qc_df,
            "By_Spike": by_spike_df,
            "Call_Counts": call_counts_df,
            "Species_Summary": species_df,
            "Detection_Calls": detection_calls_df,
            "Hit_Summary": hit_summary_df,
            "Run_Summary": summary_df,
        },
    )
    build_html_report(
        out_path=out_dir / "kmersutra_spikein_overall_summary.html",
        run_qc_df=run_qc_df,
        by_spike_df=by_spike_df,
        call_counts_df=call_counts_df,
        species_df=species_df,
        summary_df=summary_df,
        max_html_rows=max_html_rows,
    )


def main() -> None:
    """Run the command-line interface."""
    args = parse_args()
    configure_logging(verbose=args.verbose)
    expected_spike_levels = parse_expected_spike_levels(
        raw_value=args.expected_spike_levels,
    )
    run_summary_workflow(
        input_dirs=args.input_dirs,
        out_dir=Path(args.out_dir).expanduser().resolve(),
        summary_name=args.summary_name,
        run_glob=args.run_glob,
        expected_replicates=args.expected_replicates,
        expected_spike_levels=expected_spike_levels,
        max_html_rows=args.max_html_rows,
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as error:  # noqa: BLE001 - CLI should log any fatal error.
        LOGGER.error("KmerSutra summary failed: %s", error)
        sys.exit(1)
